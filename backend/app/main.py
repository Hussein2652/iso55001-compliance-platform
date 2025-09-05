import json
import os
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, List

from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from pydantic.generics import GenericModel
from typing import TypeVar, Generic
from .ai_model import ModelClient
from sqlalchemy import (
    create_engine,
    text,
    MetaData,
    Table,
    Column,
    String,
    Integer,
)
from contextlib import asynccontextmanager
from .auth import require_write_auth, role_required, optional_jwt_claims
from prometheus_client import Counter, Histogram, CONTENT_TYPE_LATEST, generate_latest
from .logging_config import get_logger
import time
from enum import Enum
import csv
import io
from io import BytesIO
from fastapi import APIRouter
import subprocess
import re
from typing import Dict
import math
import hashlib
import base64


# --- S3/MinIO client helper ---
def _get_s3_client():
    endpoint = os.getenv("OBJECT_STORE_ENDPOINT")
    bucket = os.getenv("OBJECT_STORE_BUCKET")
    access = os.getenv("OBJECT_STORE_ACCESS_KEY")
    secret = os.getenv("OBJECT_STORE_SECRET_KEY")
    if not (endpoint and bucket and access and secret):
        return None, None
    try:
        import boto3
        from botocore.config import Config as BotoConfig
        use_path = os.getenv("OBJECT_STORE_USE_PATH_STYLE", "true").lower() in ("1", "true", "yes")
        client = boto3.client(
            "s3",
            endpoint_url=endpoint,
            aws_access_key_id=access,
            aws_secret_access_key=secret,
            region_name=os.getenv("AWS_REGION", "us-east-1"),
            config=BotoConfig(signature_version="s3v4", s3={"addressing_style": "path" if use_path else "virtual"}),
        )
        return client, bucket
    except Exception:
        return None, None
import uuid
import shutil


# --- Config ---
ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_DB_URL = os.getenv("DATABASE_URL", f"sqlite:///{ROOT_DIR}/backend/app.db")
EVIDENCE_DIR = Path(os.getenv("EVIDENCE_DIR", ROOT_DIR / "backend" / "evidence")).resolve()

# SQLAlchemy engine and metadata
engine = create_engine(DEFAULT_DB_URL, future=True)
metadata = MetaData()

clauses_table = Table(
    "clauses",
    metadata,
    Column("clause_id", String, primary_key=True),
    Column("title", String, nullable=False),
    Column("summary", String, nullable=False),
)

assessments_table = Table(
    "assessments",
    metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("clause_id", String, nullable=False),
    Column("status", String, nullable=False),
    Column("evidence", String),
    Column("owner", String),
    Column("due_date", String),
    Column("created_at", String, nullable=False),
    Column("updated_at", String, nullable=False),
    Column("org_id", String),
    Column("created_by", String),
    Column("updated_by", String),
    Column("request_id", String),
)

attachments_table = Table(
    "attachments",
    metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("assessment_id", Integer, nullable=False),
    Column("filename", String, nullable=False),
    Column("content_type", String),
    Column("size", Integer),
    Column("stored_path", String, nullable=False),
    Column("created_at", String, nullable=False),
    Column("org_id", String),
    Column("created_by", String),
    Column("request_id", String),
    Column("sha256", String),
    Column("retention_hold", String),
    Column("retention_until", String),
    Column("disposition", String),
)

organizations_table = Table(
    "organizations",
    metadata,
    Column("id", String, primary_key=True),
    Column("name", String, nullable=False),
    Column("created_at", String, nullable=False),
)

# Audits and Nonconformities
audits_table = Table(
    "audits",
    metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("title", String, nullable=False),
    Column("description", String),
    Column("status", String, nullable=False),
    Column("scheduled_date", String),
    Column("completed_date", String),
    Column("created_at", String, nullable=False),
    Column("updated_at", String, nullable=False),
    Column("org_id", String),
    Column("created_by", String),
    Column("updated_by", String),
    Column("request_id", String),
)

nonconformities_table = Table(
    "nonconformities",
    metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("audit_id", Integer),
    Column("clause_id", String),
    Column("severity", String, nullable=False),
    Column("status", String, nullable=False),
    Column("state", String),
    Column("description", String, nullable=False),
    Column("corrective_action", String),
    Column("containment", String),
    Column("root_cause", String),
    Column("preventive_action", String),
    Column("verification_method", String),
    Column("verified_by", String),
    Column("verified_on", String),
    Column("owner", String),
    Column("due_date", String),
    Column("closed_date", String),
    Column("created_at", String, nullable=False),
    Column("updated_at", String, nullable=False),
    Column("org_id", String),
    Column("created_by", String),
    Column("updated_by", String),
    Column("request_id", String),
)

# Documents & KPI tables (for SQLite dev init)
documents_table = Table(
    "documents",
    metadata,
    Column("id", String, primary_key=True),
    Column("org_id", String, nullable=False),
    Column("type", String),
    Column("title", String, nullable=False),
    Column("version", String),
    Column("status", String),
    Column("approver_id", String),
    Column("effective_date", String),
    Column("next_review", String),
    Column("s3_key", String),
    Column("retention_until", String),
    Column("created_at", String, nullable=False),
    Column("updated_at", String, nullable=False),
    Column("created_by", String),
    Column("updated_by", String),
    Column("request_id", String),
)

kpi_definitions_table = Table(
    "kpi_definitions",
    metadata,
    Column("id", String, primary_key=True),
    Column("org_id", String, nullable=False),
    Column("name", String, nullable=False),
    Column("method", String),
    Column("frequency", String),
    Column("target", String),
    Column("owner_id", String),
    Column("created_at", String, nullable=False),
    Column("updated_at", String, nullable=False),
)

kpi_results_table = Table(
    "kpi_results",
    metadata,
    Column("id", String, primary_key=True),
    Column("kpi_id", String, nullable=False),
    Column("period", String, nullable=False),
    Column("value", String),
    Column("evaluation", String),
    Column("created_at", String, nullable=False),
)

management_reviews_table = Table(
    "management_reviews",
    metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("title", String, nullable=False),
    Column("period_start", String),
    Column("period_end", String),
    Column("meeting_date", String),
    Column("participants", String),
    Column("summary", String),
    Column("decisions", String),
    Column("actions", String),
    Column("created_at", String, nullable=False),
    Column("updated_at", String, nullable=False),
    Column("org_id", String),
    Column("created_by", String),
    Column("updated_by", String),
    Column("request_id", String),
)


# --- AI / RAG tables ---
ai_documents_table = Table(
    "ai_documents",
    metadata,
    Column("id", String, primary_key=True),
    Column("org_id", String, nullable=False),
    Column("source_type", String),
    Column("title", String),
    Column("path", String),
    Column("mime", String),
    Column("hash", String),
    Column("created_at", String, nullable=False),
)

ai_chunks_table = Table(
    "ai_chunks",
    metadata,
    Column("id", String, primary_key=True),
    Column("document_id", String, nullable=False),
    Column("ord", Integer, nullable=False),
    Column("text", String, nullable=False),
    Column("token_count", Integer),
)

ai_embeddings_table = Table(
    "ai_embeddings",
    metadata,
    Column("id", String, primary_key=True),
    Column("chunk_id", String, nullable=False),
    # For SQLite we store JSON-encoded vectors; for Postgres this table
    # exists with native vector type and we will insert via literal casting.
    Column("vector", String),
)

ai_runs_table = Table(
    "ai_runs",
    metadata,
    Column("id", String, primary_key=True),
    Column("org_id", String, nullable=False),
    Column("user_id", String),
    Column("task", String, nullable=False),
    Column("model", String),
    Column("temperature", String),
    Column("top_p", String),
    Column("seed", String),
    Column("inputs_json", String),
    Column("retrieved_refs_json", String),
    Column("output_json", String),
    Column("created_at", String, nullable=False),
)


def init_db():
    metadata.create_all(engine)


def seed_clauses_if_empty():
    with engine.begin() as conn:
        cnt = conn.execute(text("SELECT COUNT(1) FROM clauses")).scalar_one()
        if cnt and int(cnt) > 0:
            return
        seed_file = ROOT_DIR / "data" / "project_files" / "iso55001_clauses_seed.json"
        if not seed_file.exists():
            return
        clauses = json.loads(seed_file.read_text(encoding="utf-8"))
        for row in clauses:
            conn.execute(
                text(
                    "INSERT INTO clauses (clause_id, title, summary) VALUES (:clause_id, :title, :summary) ON CONFLICT(clause_id) DO NOTHING"
                    if DEFAULT_DB_URL.startswith("sqlite")
                    else "INSERT INTO clauses (clause_id, title, summary) VALUES (:clause_id, :title, :summary) ON CONFLICT DO NOTHING"
                ),
                {"clause_id": row["clause_id"], "title": row["title"], "summary": row["summary"]},
            )


# --- Schemas ---
class Clause(BaseModel):
    clause_id: str
    title: str
    summary: str

class StatusEnum(str, Enum):
    Compliant = "Compliant"
    Partial = "Partial"
    Noncompliant = "Noncompliant"
    NotApplicable = "NotApplicable"


class AssessmentCreate(BaseModel):
    clause_id: str = Field(..., description="ISO 55001 clause id, e.g., '4.1'")
    status: StatusEnum = Field(..., description="Compliant | Partial | Noncompliant | NotApplicable")
    evidence: Optional[str] = None
    owner: Optional[str] = None
    due_date: Optional[date] = Field(None, description="ISO date, e.g., 2025-03-31")
    model_config = {"json_schema_extra": {"examples": [{
        "clause_id": "4.1",
        "status": "Compliant",
        "evidence": "Context analysis documented in POL-001",
        "owner": "QA",
        "due_date": "2025-12-31"
    }]}}


class AssessmentUpdate(BaseModel):
    status: Optional[StatusEnum] = None
    evidence: Optional[str] = None
    owner: Optional[str] = None
    due_date: Optional[date] = None
    model_config = {"json_schema_extra": {"examples": [{"status": "Partial", "owner": "Ops"}]}}


class Assessment(BaseModel):
    id: int
    clause_id: str
    status: StatusEnum
    evidence: Optional[str]
    owner: Optional[str]
    due_date: Optional[date]
    created_at: str
    updated_at: str


class Attachment(BaseModel):
    id: int
    assessment_id: int
    filename: str
    content_type: Optional[str] = None
    size: Optional[int] = None
    created_at: str


class PresignRequest(BaseModel):
    assessment_id: int
    filename: str
    content_type: Optional[str] = None
    size: Optional[int] = None
    model_config = {"json_schema_extra": {"examples": [{"assessment_id": 1, "filename": "evidence.pdf", "content_type": "application/pdf", "size": 1048576}]}}


class PresignResponse(BaseModel):
    upload_url: str
    object_key: str
    headers: Dict[str, str] = {}


class AttachmentComplete(BaseModel):
    assessment_id: int
    object_key: str
    filename: str
    content_type: Optional[str] = None
    size: Optional[int] = None
    sha256: Optional[str] = None
    retention_hold: Optional[bool] = None
    retention_until: Optional[date] = None
    disposition: Optional[str] = None
    model_config = {"json_schema_extra": {"examples": [{"assessment_id": 1, "object_key": "orgA/assessments/1/abc123_evidence.pdf", "filename": "evidence.pdf", "content_type": "application/pdf", "size": 1048576, "sha256": "..."}]}}


# --- API helpers ---
class ErrorResponse(BaseModel):
    detail: str

T = TypeVar("T")


class Envelope(GenericModel, Generic[T]):
    items: List[T]
    total: int
    limit: int
    offset: int


class OrganizationCreate(BaseModel):
    id: str
    name: str
    model_config = {"json_schema_extra": {"examples": [{"id": "orgA", "name": "Organization A"}]}}


class OrganizationUpdate(BaseModel):
    name: Optional[str] = None
    model_config = {"json_schema_extra": {"examples": [{"name": "Organization A (Renamed)"}]}}


class Organization(BaseModel):
    id: str
    name: str
    created_at: str


# --- Documents & KPI Schemas ---
class DocumentCreate(BaseModel):
    org_id: str
    type: Optional[str] = None
    title: str
    version: Optional[str] = None
    status: Optional[str] = None
    approver_id: Optional[str] = None
    effective_date: Optional[date] = None
    next_review: Optional[date] = None
    s3_key: Optional[str] = None
    retention_until: Optional[date] = None


class Document(BaseModel):
    id: str
    org_id: str
    type: Optional[str]
    title: str
    version: Optional[str]
    status: Optional[str]
    approver_id: Optional[str]
    effective_date: Optional[date] = None
    next_review: Optional[date] = None
    s3_key: Optional[str]
    retention_until: Optional[date] = None
    created_at: str
    updated_at: str


class DocumentUpdate(BaseModel):
    title: Optional[str] = None
    version: Optional[str] = None
    status: Optional[str] = None
    approver_id: Optional[str] = None
    effective_date: Optional[date] = None
    next_review: Optional[date] = None
    s3_key: Optional[str] = None
    retention_until: Optional[date] = None


class KPIOverview(BaseModel):
    total_definitions: int
    total_results: int
    latest_periods: List[str]


# --- Planning Schemas ---
class SAMPCreate(BaseModel):
    org_id: str
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    value_definition: Optional[str] = None
    decision_criteria: Optional[str] = None
    risk_appetite: Optional[str] = None
    finance_link: Optional[str] = None


class SAMP(BaseModel):
    id: str
    org_id: str
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    value_definition: Optional[str] = None
    decision_criteria: Optional[str] = None
    risk_appetite: Optional[str] = None
    finance_link: Optional[str] = None
    created_at: str
    updated_at: str


class ObjectiveCreate(BaseModel):
    org_id: str
    samp_id: Optional[str] = None
    name: str
    measure: Optional[str] = None
    target: Optional[str] = None
    due_date: Optional[date] = None
    owner_id: Optional[str] = None
    stakeholder_ref: Optional[str] = None


class Objective(BaseModel):
    id: str
    org_id: str
    samp_id: Optional[str]
    name: str
    measure: Optional[str]
    target: Optional[str]
    due_date: Optional[date] = None
    owner_id: Optional[str] = None
    stakeholder_ref: Optional[str] = None
    created_at: str
    updated_at: str


class AuditStatus(str, Enum):
    Planned = "Planned"
    InProgress = "InProgress"
    Completed = "Completed"
    Cancelled = "Cancelled"


class AuditCreate(BaseModel):
    title: str
    description: Optional[str] = None
    status: AuditStatus = AuditStatus.Planned
    scheduled_date: Optional[date] = None
    model_config = {"json_schema_extra": {"examples": [{"title": "Internal audit Q1", "status": "Planned", "scheduled_date": "2025-01-15"}]}}


class AuditUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    status: Optional[AuditStatus] = None
    scheduled_date: Optional[date] = None
    completed_date: Optional[date] = None
    model_config = {"json_schema_extra": {"examples": [{"status": "InProgress"}]}}


class Audit(BaseModel):
    id: int
    title: str
    description: Optional[str]
    status: AuditStatus
    scheduled_date: Optional[date]
    completed_date: Optional[date]
    created_at: str
    updated_at: str


class SeverityEnum(str, Enum):
    Minor = "Minor"
    Major = "Major"
    Critical = "Critical"


class NCStatusEnum(str, Enum):
    Open = "Open"
    InProgress = "InProgress"
    Closed = "Closed"


class NCStateEnum(str, Enum):
    New = "New"
    Analysis = "Analysis"
    Action = "Action"
    Verification = "Verification"
    Closed = "Closed"


class NonconformityCreate(BaseModel):
    description: str
    severity: SeverityEnum
    status: NCStatusEnum = NCStatusEnum.Open
    state: Optional[NCStateEnum] = NCStateEnum.New
    audit_id: Optional[int] = None
    clause_id: Optional[str] = None
    corrective_action: Optional[str] = None
    containment: Optional[str] = None
    root_cause: Optional[str] = None
    preventive_action: Optional[str] = None
    verification_method: Optional[str] = None
    owner: Optional[str] = None
    due_date: Optional[date] = None
    model_config = {"json_schema_extra": {"examples": [{
        "description": "Policy not documented",
        "severity": "Major",
        "status": "Open",
        "audit_id": 1,
        "clause_id": "4.1",
        "owner": "QA",
        "due_date": "2025-02-28"
    }]}}


class NonconformityUpdate(BaseModel):
    description: Optional[str] = None
    severity: Optional[SeverityEnum] = None
    status: Optional[NCStatusEnum] = None
    state: Optional[NCStateEnum] = None
    audit_id: Optional[int] = None
    clause_id: Optional[str] = None
    corrective_action: Optional[str] = None
    containment: Optional[str] = None
    root_cause: Optional[str] = None
    preventive_action: Optional[str] = None
    verification_method: Optional[str] = None
    verified_by: Optional[str] = None
    verified_on: Optional[date] = None
    owner: Optional[str] = None
    due_date: Optional[date] = None
    closed_date: Optional[date] = None
    model_config = {"json_schema_extra": {"examples": [{"status": "Closed", "closed_date": "2025-03-01"}]}}


class Nonconformity(BaseModel):
    id: int
    description: str
    severity: SeverityEnum
    status: NCStatusEnum
    state: Optional[NCStateEnum] = None
    audit_id: Optional[int]
    clause_id: Optional[str]
    corrective_action: Optional[str]
    containment: Optional[str]
    root_cause: Optional[str]
    preventive_action: Optional[str]
    verification_method: Optional[str]
    verified_by: Optional[str]
    verified_on: Optional[date]
    owner: Optional[str]
    due_date: Optional[date]
    closed_date: Optional[date]
    created_at: str
    updated_at: str


class ManagementReviewCreate(BaseModel):
    title: str
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    meeting_date: Optional[date] = None
    participants: Optional[str] = None
    summary: Optional[str] = None
    decisions: Optional[str] = None
    actions: Optional[str] = None
    model_config = {"json_schema_extra": {"examples": [{
        "title": "Q1 Review",
        "period_start": "2025-01-01",
        "period_end": "2025-03-31",
        "meeting_date": "2025-04-10",
        "participants": "CEO, COO",
        "summary": "Good progress"
    }]}}


class ManagementReviewUpdate(BaseModel):
    title: Optional[str] = None
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    meeting_date: Optional[date] = None
    participants: Optional[str] = None
    summary: Optional[str] = None
    decisions: Optional[str] = None
    actions: Optional[str] = None
    model_config = {"json_schema_extra": {"examples": [{"decisions": "Allocate budget"}]}}


class ManagementReview(BaseModel):
    id: int
    title: str
    period_start: Optional[date]
    period_end: Optional[date]
    meeting_date: Optional[date]
    participants: Optional[str]
    summary: Optional[str]
    decisions: Optional[str]
    actions: Optional[str]
    created_at: str
    updated_at: str


# --- AI Schemas ---
class AIChunkerConfig(BaseModel):
    max_tokens: int = 800
    overlap: int = 120


class AIDocumentIn(BaseModel):
    title: Optional[str] = None
    path: Optional[str] = None
    mime: Optional[str] = None
    text: Optional[str] = None


class AIIngestRequest(BaseModel):
    org_id: str
    document: AIDocumentIn
    chunker: Optional[AIChunkerConfig] = None


class AISearchRequest(BaseModel):
    org_id: str
    query: str
    top_k: int = 6


class AISearchHit(BaseModel):
    chunk_id: str
    score: float
    doc_title: Optional[str] = None
    ref: Dict[str, str]


class AISearchResponse(BaseModel):
    hits: List[AISearchHit]


class AIGenerateIn(BaseModel):
    org_id: str
    task: str  # e.g., capa_draft, clause_map, samp_draft, objective_smartify, risk_treatment_suggest
    inputs: Dict
    retrieval: Optional[Dict] = None
    model: Optional[Dict] = None


class CAPAOut(BaseModel):
    containment: str
    root_cause: str
    corrective_action: str
    preventive_action: str
    verification_method: str
    citations: List[Dict]


class AIVisionIn(BaseModel):
    org_id: str
    image_path: str
    task: str  # nameplate_ocr|scan_layout
    hints: Optional[Dict] = None


class AIVisionOut(BaseModel):
    fields: Dict[str, str]
    citations: List[Dict]


class ReviewPackIn(BaseModel):
    org_id: str
    period_start: Optional[date] = None
    period_end: Optional[date] = None


class ReviewPackOut(BaseModel):
    document_id: str
    title: str
    narrative: str
    citations: List[Dict]


# --- App ---
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    seed_clauses_if_empty()
    # Ensure object store bucket exists (if configured)
    try:
        client, bucket = _get_s3_client()
        if client and bucket:
            buckets = [b.get('Name') for b in client.list_buckets().get('Buckets', [])]
            if bucket not in buckets:
                client.create_bucket(Bucket=bucket)
    except Exception:
        pass
    yield


openapi_tags = [
    {"name": "Health", "description": "Service status and metrics"},
    {"name": "Clauses", "description": "ISO 55001 clauses catalog"},
    {"name": "Assessments", "description": "Clause assessment records"},
    {"name": "Audits", "description": "Audit program and findings"},
    {"name": "Nonconformities", "description": "NC/CAPA workflows"},
    {"name": "Management Reviews", "description": "9.3 management review records"},
    {"name": "Attachments", "description": "Evidence uploads and links"},
    {"name": "KPIs", "description": "Performance metrics and trends"},
    {"name": "Exports", "description": "CSV/XLSX exports"},
    {"name": "Organizations", "description": "Tenant management (admin)"},
    {"name": "Setup", "description": "Admin bootstrap utilities"},
    {"name": "AI", "description": "RAG ingestion and search"},
    {"name": "Documents", "description": "Controlled documents (7.5)"},
    {"name": "KPIs", "description": "Performance indicators"},
]

app = FastAPI(
    title="ISO 55001 Compliance API",
    version="0.1.0",
    description="API for ISO 55001 readiness and compliance: clauses, assessments, audits, NCs, reviews, KPIs, and evidence.",
    lifespan=lifespan,
    openapi_tags=openapi_tags,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=(os.getenv("CORS_ALLOWED_ORIGINS", "*").split(",") if os.getenv("CORS_ALLOWED_ORIGINS") else ["*"]),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static dashboard
STATIC_DIR = ROOT_DIR / "backend" / "app" / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

@app.get("/dashboard")
def dashboard():
    index = STATIC_DIR / "index.html"
    if not index.exists():
        raise HTTPException(status_code=404, detail="Dashboard not available")
    return FileResponse(str(index))


# --- Documents CRUD (minimal) ---
@app.post("/documents", response_model=Document, tags=["Documents"])
def create_document(payload: DocumentCreate, _auth=Depends(role_required("editor"))):
    try:
        init_db()
    except Exception:
        pass
    now = datetime.utcnow().isoformat()
    did = uuid.uuid4().hex
    with engine.begin() as conn:
        conn.execute(text(
            """
            INSERT INTO documents (id, org_id, type, title, version, status, approver_id, effective_date, next_review, s3_key, retention_until, created_at, updated_at)
            VALUES (:id,:org,:type,:title,:version,:status,:approver,:effective,:next_review,:s3,:retention,:created,:updated)
            """
        ), {
            "id": did,
            "org": payload.org_id,
            "type": payload.type,
            "title": payload.title,
            "version": payload.version,
            "status": payload.status,
            "approver": payload.approver_id,
            "effective": payload.effective_date.isoformat() if payload.effective_date else None,
            "next_review": payload.next_review.isoformat() if payload.next_review else None,
            "s3": payload.s3_key,
            "retention": payload.retention_until.isoformat() if payload.retention_until else None,
            "created": now,
            "updated": now,
        })
        row = conn.execute(text("SELECT * FROM documents WHERE id = :id"), {"id": did}).mappings().first()
    d = dict(row)
    for k in ("effective_date","next_review","retention_until"):
        if d.get(k):
            try:
                d[k] = date.fromisoformat(d[k])
            except Exception:
                d[k] = None
    return Document(**d)


@app.get("/documents", response_model=List[Document], tags=["Documents"])
def list_documents(org_id: Optional[str] = None, status: Optional[str] = None, doc_type: Optional[str] = None, limit: int = 50, offset: int = 0):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    base = "SELECT * FROM documents"
    filters = []
    params = {"limit": limit, "offset": offset}
    if org_id:
        filters.append("org_id = :org_id"); params["org_id"] = org_id
    if status:
        filters.append("status = :status"); params["status"] = status
    if doc_type:
        filters.append("type = :type"); params["type"] = doc_type
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
    out = []
    for r in rows:
        d = dict(r)
        for k in ("effective_date","next_review","retention_until"):
            if d.get(k):
                try:
                    d[k] = date.fromisoformat(d[k])
                except Exception:
                    d[k] = None
        out.append(Document(**d))
    return out


@app.get("/documents/{doc_id}", response_model=Document, tags=["Documents"])
def get_document(doc_id: str):
    with engine.connect() as conn:
        row = conn.execute(text("SELECT * FROM documents WHERE id = :id"), {"id": doc_id}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Document not found")
    d = dict(row)
    for k in ("effective_date","next_review","retention_until"):
        if d.get(k):
            try:
                d[k] = date.fromisoformat(d[k])
            except Exception:
                d[k] = None
    return Document(**d)


@app.patch("/documents/{doc_id}", response_model=Document, tags=["Documents"])
def update_document(doc_id: str, payload: DocumentUpdate, _auth=Depends(role_required("editor"))):
    updates = {}
    for fld in ("title","version","status","approver_id","s3_key"):
        val = getattr(payload, fld)
        if val is not None:
            updates[fld] = val
    for fld in ("effective_date","next_review","retention_until"):
        val = getattr(payload, fld)
        if val is not None:
            updates[fld] = val.isoformat() if val else None
    # enforce status transitions
    if "status" in updates:
        with engine.connect() as conn:
            cur = conn.execute(text("SELECT status FROM documents WHERE id = :id"), {"id": doc_id}).mappings().first()
        curr = (cur or {}).get("status")
        allowed = {
            None: {"Draft"},
            "Draft": {"Approved","Obsolete"},
            "Approved": {"Obsolete"},
            "Obsolete": set(),
        }
        nxt = updates["status"]
        if curr not in allowed or nxt not in allowed[curr]:
            raise HTTPException(status_code=400, detail=f"Invalid status transition: {curr} -> {nxt}")
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.utcnow().isoformat()
    updates["id"] = doc_id
    set_clause = ", ".join([f"{k} = :{k}" for k in updates.keys() if k != "id"])
    with engine.begin() as conn:
        res = conn.execute(text(f"UPDATE documents SET {set_clause} WHERE id = :id"), updates)
        if res.rowcount == 0:
            raise HTTPException(status_code=404, detail="Document not found")
        row = conn.execute(text("SELECT * FROM documents WHERE id = :id"), {"id": doc_id}).mappings().first()
    d = dict(row)
    for k in ("effective_date","next_review","retention_until"):
        if d.get(k):
            try:
                d[k] = date.fromisoformat(d[k])
            except Exception:
                d[k] = None
    return Document(**d)
@app.get("/kpi/overview", response_model=KPIOverview, tags=["KPIs"])
def kpi_overview(org_id: str):
    try:
        init_db()
    except Exception:
        pass
    with engine.connect() as conn:
        total_def = conn.execute(text("SELECT COUNT(*) FROM kpi_definitions WHERE org_id = :org"), {"org": org_id}).scalar_one()
        total_res = conn.execute(text("SELECT COUNT(*) FROM kpi_results kr JOIN kpi_definitions kd ON kd.id = kr.kpi_id WHERE kd.org_id = :org"), {"org": org_id}).scalar_one()
        periods = conn.execute(text("SELECT DISTINCT period FROM kpi_results kr JOIN kpi_definitions kd ON kd.id = kr.kpi_id WHERE kd.org_id = :org ORDER BY period DESC LIMIT 6"), {"org": org_id}).fetchall()
    latest_periods = [p[0] for p in periods]
    return KPIOverview(total_definitions=int(total_def or 0), total_results=int(total_res or 0), latest_periods=latest_periods)


# --- Planning minimal CRUD ---
@app.post("/samps", response_model=SAMP, tags=["Planning"])
def create_samp(payload: SAMPCreate, _auth=Depends(role_required("editor"))):
    now = datetime.utcnow().isoformat()
    sid = uuid.uuid4().hex
    with engine.begin() as conn:
        conn.execute(text("INSERT INTO samps (id, org_id, period_start, period_end, value_definition, decision_criteria, risk_appetite, finance_link, created_at, updated_at) VALUES (:id,:org,:ps,:pe,:vd,:dc,:ra,:fl,:c,:u)"), {
            "id": sid, "org": payload.org_id,
            "ps": payload.period_start.isoformat() if payload.period_start else None,
            "pe": payload.period_end.isoformat() if payload.period_end else None,
            "vd": payload.value_definition, "dc": payload.decision_criteria,
            "ra": payload.risk_appetite, "fl": payload.finance_link,
            "c": now, "u": now,
        })
        row = conn.execute(text("SELECT * FROM samps WHERE id = :id"), {"id": sid}).mappings().first()
    d = dict(row)
    for k in ("period_start","period_end"):
        if d.get(k):
            try:
                d[k] = date.fromisoformat(d[k])
            except Exception:
                d[k] = None
    return SAMP(**d)


@app.get("/samps", response_model=List[SAMP], tags=["Planning"])
def list_samps(org_id: Optional[str] = None, limit: int = 50, offset: int = 0):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    base = "SELECT * FROM samps"
    params = {"limit": limit, "offset": offset}
    if org_id:
        base += " WHERE org_id = :org"; params["org"] = org_id
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
    out = []
    for r in rows:
        d = dict(r)
        for k in ("period_start","period_end"):
            if d.get(k):
                try:
                    d[k] = date.fromisoformat(d[k])
                except Exception:
                    d[k] = None
        out.append(SAMP(**d))
    return out


@app.post("/objectives", response_model=Objective, tags=["Planning"])
def create_objective(payload: ObjectiveCreate, _auth=Depends(role_required("editor"))):
    now = datetime.utcnow().isoformat()
    oid = uuid.uuid4().hex
    with engine.begin() as conn:
        conn.execute(text("INSERT INTO am_objectives (id, org_id, samp_id, name, measure, target, due_date, owner_id, stakeholder_ref, created_at, updated_at) VALUES (:id,:org,:samp,:name,:measure,:target,:due,:owner,:stake,:c,:u)"), {
            "id": oid, "org": payload.org_id, "samp": payload.samp_id,
            "name": payload.name, "measure": payload.measure, "target": payload.target,
            "due": payload.due_date.isoformat() if payload.due_date else None,
            "owner": payload.owner_id, "stake": payload.stakeholder_ref,
            "c": now, "u": now,
        })
        row = conn.execute(text("SELECT * FROM am_objectives WHERE id = :id"), {"id": oid}).mappings().first()
    d = dict(row)
    if d.get("due_date"):
        try:
            d["due_date"] = date.fromisoformat(d["due_date"])  # type: ignore
        except Exception:
            d["due_date"] = None
    return Objective(**d)


@app.get("/objectives", response_model=List[Objective], tags=["Planning"])
def list_objectives(org_id: Optional[str] = None, samp_id: Optional[str] = None, limit: int = 50, offset: int = 0):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    base = "SELECT * FROM am_objectives"
    filters = []
    params = {"limit": limit, "offset": offset}
    if org_id:
        filters.append("org_id = :org"); params["org"] = org_id
    if samp_id:
        filters.append("samp_id = :samp"); params["samp"] = samp_id
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
    out = []
    for r in rows:
        d = dict(r)
        if d.get("due_date"):
            try:
                d["due_date"] = date.fromisoformat(d["due_date"])  # type: ignore
            except Exception:
                d["due_date"] = None
        out.append(Objective(**d))
    return out


# --- AI utilities (chunking, embeddings) ---
def _tokenize(text: str) -> List[str]:
    # simple whitespace tokenization
    return re.findall(r"\S+", text or "")


def _chunk_text(text: str, max_tokens: int = 800, overlap: int = 120) -> List[str]:
    tokens = _tokenize(text)
    if max_tokens <= 0:
        max_tokens = 800
    if overlap < 0:
        overlap = 0
    chunks: List[str] = []
    i = 0
    n = len(tokens)
    step = max(1, max_tokens - overlap)
    while i < n:
        window = tokens[i : min(n, i + max_tokens)]
        if not window:
            break
        chunks.append(" ".join(window))
        i += step
    return chunks


def _hash_embedding(text: str, dim: int = 1024) -> List[float]:
    # Deterministic, lightweight fallback embedding (not semantic, dev-only)
    if dim <= 0:
        dim = 1024
    vec = [0.0] * dim
    for tok in _tokenize(text.lower()):
        h = hashlib.sha256(tok.encode("utf-8")).digest()
        # use first 8*4 bytes as 32-bit ints
        for j in range(0, min(dim, len(h))):
            vec[j] += h[j] / 255.0
    # L2 normalize
    norm = math.sqrt(sum(v * v for v in vec)) or 1.0
    return [v / norm for v in vec]


def _cosine(a: List[float], b: List[float]) -> float:
    if not a or not b:
        return 0.0
    n = min(len(a), len(b))
    dot = sum(a[i] * b[i] for i in range(n))
    na = math.sqrt(sum(a[i] * a[i] for i in range(n))) or 1.0
    nb = math.sqrt(sum(b[i] * b[i] for i in range(n))) or 1.0
    return dot / (na * nb)


def _vector_to_pg_text(vec: List[float]) -> str:
    # pgvector accepts "[v1, v2, ...]" textual input for vector columns
    return "[" + ",".join(f"{x:.6f}" for x in vec) + "]"


_PII_EMAIL = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
_PII_PHONE = re.compile(r"\+?\d[\d\-()\s]{6,}\d")


def _redact_pii(text: str) -> str:
    if not text:
        return text
    t = _PII_EMAIL.sub("<redacted:email>", text)
    t = _PII_PHONE.sub("<redacted:phone>", t)
    return t


def _redact_dict(d: Dict) -> Dict:
    try:
        js = json.dumps(d)
        return json.loads(_redact_pii(js))
    except Exception:
        return d


def _retrieve_context(org_id: str, query: str, top_k: int = 6):
    qvec = _hash_embedding(query, dim=1024)
    with engine.connect() as conn:
        sql = text(
            """
            SELECT e.id as eid, e.chunk_id as chunk_id, e.vector as vector, c.ord as ord, c.text as chunk_text,
                   d.title as title, d.id as document_id
            FROM ai_embeddings e
            JOIN ai_chunks c ON c.id = e.chunk_id
            JOIN ai_documents d ON d.id = c.document_id
            WHERE d.org_id = :org_id
            """
        )
        rows = conn.execute(sql, {"org_id": org_id}).mappings().all()
    scored = []
    for r in rows:
        vec_raw = r.get("vector")
        vec: List[float] = []
        if isinstance(vec_raw, str):
            try:
                vec = json.loads(vec_raw)
            except Exception:
                s = vec_raw.strip()
                if s.startswith("[") and s.endswith("]"):
                    try:
                        vec = [float(x) for x in s[1:-1].split(",") if x.strip()]
                    except Exception:
                        vec = []
        if not vec:
            continue
        score = _cosine(qvec, vec)
        scored.append((score, r))
    scored.sort(key=lambda x: x[0], reverse=True)
    out = []
    for score, r in scored[:top_k]:
        out.append({
            "score": float(round(score, 6)),
            "chunk_id": r.get("chunk_id"),
            "ord": int(r.get("ord") or 0),
            "text": r.get("chunk_text"),
            "document_id": r.get("document_id"),
            "title": r.get("title"),
        })
    return out


def _read_text_from_path(path: str) -> Optional[str]:
    try:
        if not path:
            return None
        if path.startswith("s3://"):
            client, bucket = _get_s3_client()
            if not client or not bucket:
                return None
            # path like s3://bucket/key or s3://key (assume configured bucket)
            p = path[5:]
            if "/" in p:
                bkt, key = p.split("/", 1)
                if bucket and bkt and bkt != bucket:
                    # override bucket when fully qualified
                    bucket = bkt
            else:
                key = p
            obj = client.get_object(Bucket=bucket, Key=key)
            data = obj["Body"].read()
            # naive: try utf-8 decode; for PDFs we'd need proper parser
            try:
                return data.decode("utf-8")
            except Exception:
                return None
        else:
            fp = Path(path)
            if fp.exists() and fp.is_file():
                try:
                    return fp.read_text(encoding="utf-8")
                except Exception:
                    return None
    except Exception:
        return None
    return None


# --- AI Endpoints ---
@app.post("/ai/ingest", tags=["AI"])
def ai_ingest(payload: AIIngestRequest, request: Request, _auth=Depends(role_required("editor"))):
    # Ensure tables exist (especially in test/dev where lifespan may not run)
    try:
        init_db()
    except Exception:
        pass
    org_id = payload.org_id
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id is required")
    # Read text content
    text_content: Optional[str] = payload.document.text if payload.document else None
    if (not text_content) and payload.document and payload.document.path:
        text_content = _read_text_from_path(payload.document.path)
    if not text_content:
        raise HTTPException(status_code=400, detail="No text content available (provide document.text or a readable path)")

    chunk_cfg = payload.chunker or AIChunkerConfig()
    chunks = _chunk_text(text_content, max_tokens=chunk_cfg.max_tokens, overlap=chunk_cfg.overlap)
    now = datetime.utcnow().isoformat()
    doc_id = uuid.uuid4().hex
    # Hash of content (simple sha256 of entire text)
    doc_hash = hashlib.sha256(text_content.encode("utf-8")).hexdigest()
    title = (payload.document.title if payload.document else None) or (Path(payload.document.path).name if payload.document and payload.document.path else None) or "document"
    mime = payload.document.mime if payload.document else None
    src_type = "upload"
    with engine.begin() as conn:
        conn.execute(
            text(
                "INSERT INTO ai_documents (id, org_id, source_type, title, path, mime, hash, created_at) VALUES (:id, :org_id, :source_type, :title, :path, :mime, :hash, :created_at)"
            ),
            {
                "id": doc_id,
                "org_id": org_id,
                "source_type": src_type,
                "title": title,
                "path": (payload.document.path if payload.document else None),
                "mime": mime,
                "hash": doc_hash,
                "created_at": now,
            },
        )

        # Determine backend
        is_sqlite = DEFAULT_DB_URL.startswith("sqlite")
        is_postgres = (not is_sqlite) and (conn.engine.dialect.name == "postgresql")
        for idx, chunk_text in enumerate(chunks):
            chunk_id = uuid.uuid4().hex
            tok_count = len(_tokenize(chunk_text))
            conn.execute(
                text(
                    "INSERT INTO ai_chunks (id, document_id, ord, text, token_count) VALUES (:id, :document_id, :ord, :text, :token_count)"
                ),
                {
                    "id": chunk_id,
                    "document_id": doc_id,
                    "ord": idx,
                    "text": chunk_text,
                    "token_count": tok_count,
                },
            )
            emb = _hash_embedding(chunk_text, dim=1024)
            if is_postgres:
                vec_txt = _vector_to_pg_text(emb)
                conn.execute(
                    text(
                        "INSERT INTO ai_embeddings (id, chunk_id, vector) VALUES (:id, :chunk_id, :vec::vector)"
                    ),
                    {"id": uuid.uuid4().hex, "chunk_id": chunk_id, "vec": vec_txt},
                )
            else:
                conn.execute(
                    text("INSERT INTO ai_embeddings (id, chunk_id, vector) VALUES (:id, :chunk_id, :vector)"),
                    {
                        "id": uuid.uuid4().hex,
                        "chunk_id": chunk_id,
                        "vector": json.dumps(emb),
                    },
                )

    return {"document_id": doc_id, "chunks": len(chunks)}


@app.post("/ai/search", response_model=AISearchResponse, tags=["AI"])
def ai_search(payload: AISearchRequest, request: Request, _auth=Depends(role_required("viewer"))):
    try:
        init_db()
    except Exception:
        pass
    org_id = payload.org_id
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id is required")
    qvec = _hash_embedding(payload.query, dim=1024)
    top_k = max(1, min(50, int(payload.top_k or 6)))
    with engine.connect() as conn:
        is_sqlite = DEFAULT_DB_URL.startswith("sqlite")
        is_postgres = (not is_sqlite) and (conn.engine.dialect.name == "postgresql")
        if is_postgres:
            sql = text(
                """
                SELECT e.id as eid, e.chunk_id as chunk_id, (e.vector)::text as vector,
                       c.ord as ord, c.text as chunk_text, d.title as title, d.id as document_id
                FROM ai_embeddings e
                JOIN ai_chunks c ON c.id = e.chunk_id
                JOIN ai_documents d ON d.id = c.document_id
                WHERE d.org_id = :org_id
                """
            )
        else:
            sql = text(
                """
                SELECT e.id as eid, e.chunk_id as chunk_id, e.vector as vector,
                       c.ord as ord, c.text as chunk_text, d.title as title, d.id as document_id
                FROM ai_embeddings e
                JOIN ai_chunks c ON c.id = e.chunk_id
                JOIN ai_documents d ON d.id = c.document_id
                WHERE d.org_id = :org_id
                """
            )
        rows = conn.execute(sql, {"org_id": org_id}).mappings().all()

    scored: List[tuple] = []
    for r in rows:
        vec_raw = r.get("vector")
        vec: List[float] = []
        if isinstance(vec_raw, str):
            # try JSON first
            try:
                vec = json.loads(vec_raw)
            except Exception:
                # try pgvector text format: [v1, v2, ...]
                s = vec_raw.strip()
                if s.startswith("[") and s.endswith("]"):
                    try:
                        vec = [float(x) for x in s[1:-1].split(",") if x.strip()]
                    except Exception:
                        vec = []
        if not vec:
            continue
        score = _cosine(qvec, vec)
        scored.append((score, r))
    scored.sort(key=lambda x: x[0], reverse=True)
    hits: List[AISearchHit] = []
    for score, r in scored[:top_k]:
        hits.append(
            AISearchHit(
                chunk_id=r.get("chunk_id"),
                score=float(round(score, 6)),
                doc_title=r.get("title"),
                ref={"document_id": r.get("document_id"), "ord": str(r.get("ord"))},
            )
        )
    return AISearchResponse(hits=hits)


@app.post("/ai/generate", tags=["AI"])
def ai_generate(payload: AIGenerateIn, request: Request, _auth=Depends(role_required("editor"))):
    # Strict tasks supported for now
    task = (payload.task or "").strip().lower()
    if task not in {"capa_draft", "clause_map", "samp_draft", "objective_smartify", "risk_treatment_suggest"}:
        raise HTTPException(status_code=400, detail="Unsupported task")
    org_id = payload.org_id
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id is required")
    retrieval = payload.retrieval or {}
    query = retrieval.get("query") or (payload.inputs.get("incident_text") if task == "capa_draft" else payload.inputs.get("text") or "")
    top_k = int(retrieval.get("top_k") or 6)
    ctx = _retrieve_context(org_id, query, top_k=top_k) if query else []
    citations = [{"document_id": c["document_id"], "ord": c["ord"]} for c in ctx]

    # Simple deterministic draft to satisfy schema + guardrails
    if task == "capa_draft":
        fields = payload.inputs or {}
        temp = float((payload.model or {}).get("temperature", 0.3) if payload.model else 0.3)
        max_new = int((payload.model or {}).get("max_new_tokens", 512) if payload.model else 512)
        required = fields.get("required_fields") or ["containment","root_cause","corrective_action","preventive_action","verification_method"]
        incident_text = str(fields.get("incident_text") or "")
        # Prepare system + user prompts
        ctx_text = "\n\n".join([f"[ref {i}] {c['text']}" for i, c in enumerate(ctx, start=1)])
        sys_p = (
            "You extract CAPA fields as strict JSON. "
            "Use only provided context and inputs. Never hallucinate. "
            "Always include a 'citations' array with {document_id, ord}."
        )
        user_p = (
            f"Context:\n{ctx_text}\n\n"
            f"Inputs:\nincident_text: {incident_text}\nrequired_fields: {required}\n\n"
            "Return JSON with keys: containment, root_cause, corrective_action, preventive_action, verification_method, citations."
        )
        client = ModelClient()
        gen = client.generate_structured(sys_p, user_p, temperature=temp, max_new_tokens=max_new)
        out = None
        if gen:
            try:
                out = CAPAOut(**gen).model_dump()
            except Exception:
                out = None
        if out is None:
            # fallback stub
            def _stub(label):
                return f"Draft {label.replace('_',' ')} based on incident context."
            out = {
                "containment": fields.get("containment") or _stub("containment"),
                "root_cause": fields.get("root_cause") or _stub("root_cause"),
                "corrective_action": fields.get("corrective_action") or _stub("corrective_action"),
                "preventive_action": fields.get("preventive_action") or _stub("preventive_action"),
                "verification_method": fields.get("verification_method") or _stub("verification_method"),
                "citations": citations[:max(1, min(6, len(citations)))] if citations else [],
            }
        # Enforce citations policy
        if not out.get("citations"):
            return {"needs_more_evidence": True, "message": "Insufficient context for grounded answer"}
        # Log ai_runs
        with engine.begin() as conn:
            conn.execute(text("INSERT INTO ai_runs (id, org_id, user_id, task, model, temperature, top_p, seed, inputs_json, retrieved_refs_json, output_json, created_at) VALUES (:id,:org,:user,:task,:model,:temp,:top_p,:seed,:inputs,:refs,:out,:ts)"), {
                "id": uuid.uuid4().hex,
                "org": org_id,
                "user": (_auth or {}).get("sub"),
                "task": task,
                "model": (payload.model or {}).get("name") if payload.model else os.getenv('LLM_HTTP_MODEL'),
                "temp": str(temp),
                "top_p": str((payload.model or {}).get("top_p")) if payload.model else None,
                "seed": str((payload.model or {}).get("seed")) if payload.model else None,
                "inputs": json.dumps(_redact_dict(payload.inputs or {})),
                "refs": json.dumps(citations),
                "out": json.dumps(out),
                "ts": datetime.utcnow().isoformat(),
            })
        return CAPAOut(**out)

    if task == "clause_map":
        text_in = (payload.inputs or {}).get("text") or ""
        # naive clause heuristic + retrieval titles
        clauses = []
        with engine.connect() as conn:
            rows = conn.execute(text("SELECT clause_id, title FROM clauses ORDER BY clause_id")).mappings().all()
        for r in rows:
            cid = r["clause_id"]
            title = r["title"]
            if cid in text_in or title.lower() in text_in.lower():
                clauses.append({"clause_id": cid, "title": title})
        if not clauses and ctx:
            clauses = [{"clause_id": c.get("title","unknown"), "title": c.get("title") or ""} for c in ctx[:3]]
        if not citations:
            return {"needs_more_evidence": True, "message": "Insufficient context for grounded mapping"}
        with engine.begin() as conn:
            conn.execute(text("INSERT INTO ai_runs (id, org_id, user_id, task, inputs_json, retrieved_refs_json, output_json, created_at) VALUES (:id,:org,:user,:task,:inputs,:refs,:out,:ts)"), {
                "id": uuid.uuid4().hex,
                "org": org_id,
                "user": (_auth or {}).get("sub"),
                "task": task,
                "inputs": json.dumps(_redact_dict(payload.inputs or {})),
                "refs": json.dumps(citations),
                "out": json.dumps({"clauses": clauses, "citations": citations}),
                "ts": datetime.utcnow().isoformat(),
            })
        return {"clauses": clauses, "citations": citations}

    if task == "samp_draft":
        fields = payload.inputs or {}
        # Build a structured plan stub
        out = {
            "value_definition": fields.get("value_definition") or "Draft value definition based on context",
            "decision_criteria": fields.get("decision_criteria") or "Draft decision criteria",
            "risk_appetite": fields.get("risk_appetite") or "Draft risk appetite",
        }
        if not citations:
            return {"needs_more_evidence": True, "message": "Insufficient context", "citations": []}
        with engine.begin() as conn:
            conn.execute(text("INSERT INTO ai_runs (id, org_id, user_id, task, inputs_json, retrieved_refs_json, output_json, created_at) VALUES (:id,:org,:user,:task,:inputs,:refs,:out,:ts)"), {
                "id": uuid.uuid4().hex,
                "org": org_id,
                "user": (_auth or {}).get("sub"),
                "task": task,
                "inputs": json.dumps(_redact_dict(payload.inputs or {})),
                "refs": json.dumps(citations),
                "out": json.dumps(out),
                "ts": datetime.utcnow().isoformat(),
            })
        out["citations"] = citations
        return out

    if task == "objective_smartify":
        fields = payload.inputs or {}
        name = fields.get("name") or "Objective"
        measure = fields.get("measure") or "Define measure"
        target = fields.get("target") or "Define target"
        out = {"name": name, "measure": measure, "target": target}
        if not citations:
            return {"needs_more_evidence": True, "message": "Insufficient context", "citations": []}
        with engine.begin() as conn:
            conn.execute(text("INSERT INTO ai_runs (id, org_id, user_id, task, inputs_json, retrieved_refs_json, output_json, created_at) VALUES (:id,:org,:user,:task,:inputs,:refs,:out,:ts)"), {
                "id": uuid.uuid4().hex,
                "org": org_id,
                "user": (_auth or {}).get("sub"),
                "task": task,
                "inputs": json.dumps(_redact_dict(payload.inputs or {})),
                "refs": json.dumps(citations),
                "out": json.dumps(out),
                "ts": datetime.utcnow().isoformat(),
            })
        out["citations"] = citations
        return out

    if task == "risk_treatment_suggest":
        fields = payload.inputs or {}
        out = {
            "treatments": [
                {"action": "Mitigation plan", "rationale": "Based on retrieved context"},
                {"action": "Monitoring", "rationale": "Track KPI and incidents"},
            ]
        }
        if not citations:
            return {"needs_more_evidence": True, "message": "Insufficient context", "citations": []}
        with engine.begin() as conn:
            conn.execute(text("INSERT INTO ai_runs (id, org_id, user_id, task, inputs_json, retrieved_refs_json, output_json, created_at) VALUES (:id,:org,:user,:task,:inputs,:refs,:out,:ts)"), {
                "id": uuid.uuid4().hex,
                "org": org_id,
                "user": (_auth or {}).get("sub"),
                "task": task,
                "inputs": json.dumps(_redact_dict(payload.inputs or {})),
                "refs": json.dumps(citations),
                "out": json.dumps(out),
                "ts": datetime.utcnow().isoformat(),
            })
        out["citations"] = citations
        return out


@app.post("/ai/vision", response_model=AIVisionOut, tags=["AI"])
def ai_vision(payload: AIVisionIn, request: Request, _auth=Depends(role_required("editor"))):
    try:
        init_db()
    except Exception:
        pass
    org_id = payload.org_id
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id is required")
    # Try VLM client
    client = ModelClient()
    vlm_out = client.vision_extract(payload.image_path, instruction=f"Task: {payload.task}. Extract fields as JSON.", hints=payload.hints)
    fields = {}
    if isinstance(vlm_out, dict):
        fields = vlm_out.get('fields', vlm_out)
    # Heuristic enrichments from filename + hints
    basename = Path(payload.image_path).name
    m = re.search(r"([A-Za-z]-\d{3,})", basename)
    if m and 'asset_tag' not in fields:
        fields['asset_tag'] = m.group(1)
    if payload.hints:
        for k, v in payload.hints.items():
            fields.setdefault(k, str(v))
    citations: List[Dict] = []
    # Log run
    with engine.begin() as conn:
        conn.execute(text("INSERT INTO ai_runs (id, org_id, user_id, task, inputs_json, output_json, created_at) VALUES (:id,:org,:user,:task,:inputs,:out,:ts)"), {
            "id": uuid.uuid4().hex,
            "org": org_id,
            "user": (_auth or {}).get("sub"),
            "task": f"vision:{payload.task}",
            "inputs": json.dumps(_redact_dict({"image_path": payload.image_path, "hints": payload.hints or {}})),
            "out": json.dumps({"fields": fields, "citations": citations}),
            "ts": datetime.utcnow().isoformat(),
        })
    return AIVisionOut(fields=fields, citations=citations)


@app.post("/ai/review-pack", response_model=ReviewPackOut, tags=["AI"])
def ai_review_pack(payload: ReviewPackIn, request: Request, _auth=Depends(role_required("editor"))):
    try:
        init_db()
    except Exception:
        pass
    org_id = payload.org_id
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id is required")
    ps = payload.period_start.isoformat() if payload.period_start else None
    pe = payload.period_end.isoformat() if payload.period_end else None
    # Collect counts for narrative
    with engine.connect() as conn:
        audits = conn.execute(text("SELECT COUNT(*) FROM audits WHERE org_id = :org" + (" AND created_at >= :ps" if ps else "") + (" AND created_at <= :pe" if pe else "")), {k:v for k,v in {"org":org_id, "ps":ps, "pe":pe}.items() if v is not None}).scalar_one()
        ncs_open = conn.execute(text("SELECT COUNT(*) FROM nonconformities WHERE org_id = :org AND status = 'Open'"), {"org": org_id}).scalar_one()
        ncs_closed = conn.execute(text("SELECT COUNT(*) FROM nonconformities WHERE org_id = :org AND status = 'Closed'"), {"org": org_id}).scalar_one()
        kpis = conn.execute(text("SELECT COUNT(*) FROM kpi_definitions WHERE org_id = :org"), {"org": org_id}).scalar_one()
    title = f"Management Review {ps or ''} to {pe or ''}".strip()
    narrative = (
        f"This period recorded {int(audits or 0)} audits. Nonconformities: "
        f"{int(ncs_open or 0)} open, {int(ncs_closed or 0)} closed. "
        f"Configured KPIs: {int(kpis or 0)}."
    )
    citations: List[Dict] = []
    doc_id = uuid.uuid4().hex
    now = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        conn.execute(text("INSERT INTO documents (id, org_id, type, title, status, created_at, updated_at) VALUES (:id,:org,:type,:title,:status,:created,:updated)"), {
            "id": doc_id,
            "org": org_id,
            "type": "review_pack",
            "title": title or "Management Review",
            "status": "Draft",
            "created": now,
            "updated": now,
        })
        conn.execute(text("INSERT INTO ai_runs (id, org_id, user_id, task, inputs_json, output_json, created_at) VALUES (:id,:org,:user,:task,:inputs,:out,:ts)"), {
            "id": uuid.uuid4().hex,
            "org": org_id,
            "user": (_auth or {}).get("sub"),
            "task": "review_pack",
            "inputs": json.dumps(_redact_dict({"period_start": ps, "period_end": pe})),
            "out": json.dumps({"document_id": doc_id, "title": title, "narrative": narrative, "citations": citations}),
            "ts": now,
        })
    return ReviewPackOut(document_id=doc_id, title=title or "Management Review", narrative=narrative, citations=citations)



# --- OpenTelemetry (optional) ---
if os.getenv("OTEL_ENABLED", "false").lower() in ("1", "true", "yes"):
    try:
        from opentelemetry import trace
        from opentelemetry.sdk.resources import Resource
        from opentelemetry.sdk.trace import TracerProvider
        from opentelemetry.sdk.trace.export import BatchSpanProcessor
        from opentelemetry.exporter.otlp.proto.http.trace_exporter import OTLPSpanExporter
        from opentelemetry.instrumentation.fastapi import FastAPIInstrumentor
        from opentelemetry.instrumentation.sqlalchemy import SQLAlchemyInstrumentor

        resource = Resource(attributes={"service.name": "iso55001-backend"})
        provider = TracerProvider(resource=resource)
        endpoint = os.getenv("OTEL_EXPORTER_OTLP_ENDPOINT")
        exporter = OTLPSpanExporter(endpoint=endpoint) if endpoint else OTLPSpanExporter()
        provider.add_span_processor(BatchSpanProcessor(exporter))
        trace.set_tracer_provider(provider)
        FastAPIInstrumentor.instrument_app(app)
        SQLAlchemyInstrumentor().instrument(engine=engine)
    except Exception:
        pass

# --- Request ID + metrics middleware ---
if 'REQUEST_COUNT' not in globals():
    REQUEST_COUNT = Counter(
        "http_requests_total",
        "Total HTTP requests",
        labelnames=("method", "status"),
    )
if 'ERROR_COUNT' not in globals():
    ERROR_COUNT = Counter(
        "http_requests_errors_total",
        "Total HTTP error responses",
        labelnames=("method", "status"),
    )

# duration histogram (seconds) with per-route label
if 'REQUEST_DURATION' not in globals():
    REQUEST_DURATION = Histogram(
        "http_request_duration_seconds",
        "HTTP request duration in seconds",
        labelnames=("method", "route", "status"),
        buckets=(0.005, 0.01, 0.025, 0.05, 0.1, 0.25, 0.5, 1.0, 2.5, 5.0, 10.0),
    )


@app.middleware("http")
async def add_request_id_and_collect_metrics(request: Request, call_next):
    req_id = request.headers.get("X-Request-ID") or uuid.uuid4().hex
    logger = get_logger()
    start = time.perf_counter()
    # Capture auth/tenant context from headers if present
    request.state.request_id = req_id
    env_mode = os.getenv("ENV", "dev").lower()
    org_header = request.headers.get("X-Org-ID")
    # Load org from JWT claims when available
    claims = optional_jwt_claims(request.headers.get("authorization"))
    claim_org = None
    if claims:
        for k in ("org", "org_id", "tenant", "tid"):
            if k in claims:
                claim_org = str(claims[k]); break
    if env_mode == "prod":
        request.state.org_id = claim_org or os.getenv("DEFAULT_ORG_ID")
    else:
        request.state.org_id = org_header or claim_org or os.getenv("DEFAULT_ORG_ID")
    try:
        response: Response = await call_next(request)
    except Exception as exc:
        duration_ms = (time.perf_counter() - start) * 1000.0
        REQUEST_COUNT.labels(method=request.method, status=str(500)).inc()
        ERROR_COUNT.labels(method=request.method, status=str(500)).inc()
        route_label = getattr(request.scope.get("route"), "path", request.url.path)
        REQUEST_DURATION.labels(method=request.method, route=route_label, status="500").observe(duration_ms / 1000.0)
        logger.error(
            __import__("json").dumps(
                {
                    "event": "request_error",
                    "request_id": req_id,
                    "method": request.method,
                    "path": request.url.path,
                    "query": request.url.query,
                    "status": 500,
                    "duration_ms": round(duration_ms, 2),
                    "client_ip": getattr(request.client, "host", None),
                    "user_agent": request.headers.get("user-agent"),
                    "error": str(exc),
                }
            )
        )
        raise
    response.headers["X-Request-ID"] = req_id
    REQUEST_COUNT.labels(method=request.method, status=str(response.status_code)).inc()
    if response.status_code >= 400:
        ERROR_COUNT.labels(method=request.method, status=str(response.status_code)).inc()
    duration_ms = (time.perf_counter() - start) * 1000.0
    route_label = getattr(request.scope.get("route"), "path", request.url.path)
    REQUEST_DURATION.labels(method=request.method, route=route_label, status=str(response.status_code)).observe(duration_ms / 1000.0)
    logger.info(
        __import__("json").dumps(
            {
                "event": "request",
                "request_id": req_id,
                "method": request.method,
                "path": request.url.path,
                "query": request.url.query,
                "status": response.status_code,
                "duration_ms": round(duration_ms, 2),
                "client_ip": getattr(request.client, "host", None),
                "user_agent": request.headers.get("user-agent"),
            }
        )
    )
    return response


# Security headers
@app.middleware("http")
async def security_headers(request: Request, call_next):
    response: Response = await call_next(request)
    if os.getenv("ENV", "dev").lower() == "prod":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["X-Content-Type-Options"] = "nosniff"
    # Basic CSP for API; adjust as needed
    response.headers["Content-Security-Policy"] = "default-src 'none'; frame-ancestors 'none'"
    return response


# Auth dependency lives in app.auth


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/metrics", tags=["Health"])
def metrics():
    return Response(generate_latest(), media_type=CONTENT_TYPE_LATEST)


# --- KPI / Reporting ---
@app.get("/kpi/overview")
def kpi_overview(
    created_from: Optional[date] = None,
    created_to: Optional[date] = None,
    as_of: Optional[date] = None,
):
    today = as_of or datetime.utcnow().date()
    def range_clause(table: str) -> tuple[str, dict]:
        clauses = []
        params: dict = {}
        if created_from:
            clauses.append(f"{table}.created_at >= :created_from")
            params["created_from"] = created_from.isoformat()
        if created_to:
            clauses.append(f"{table}.created_at <= :created_to")
            params["created_to"] = created_to.isoformat()
        return (" WHERE " + " AND ".join(clauses)) if clauses else "", params

    with engine.connect() as conn:
        a_where, a_params = range_clause("audits")
        audits_by_status = conn.execute(text(f"SELECT status, COUNT(*) as c FROM audits{a_where} GROUP BY status"), a_params).all()
        n_where, n_params = range_clause("nonconformities")
        ncs_by_status = conn.execute(text(f"SELECT status, COUNT(*) as c FROM nonconformities{n_where} GROUP BY status"), n_params).all()
        ncs_by_severity = conn.execute(text(f"SELECT severity, COUNT(*) as c FROM nonconformities{n_where} GROUP BY severity"), n_params).all()
        am_where, am_params = range_clause("assessments")
        assessments_by_status = conn.execute(text(f"SELECT status, COUNT(*) as c FROM assessments{am_where} GROUP BY status"), am_params).all()
        overdue = conn.execute(
            text("SELECT COUNT(*) FROM nonconformities WHERE status != 'Closed' AND due_date IS NOT NULL AND due_date < :today"),
            {"today": today.isoformat()},
        ).scalar_one()
        total_audits = conn.execute(text(f"SELECT COUNT(*) FROM audits{a_where}"), a_params).scalar_one()
        total_ncs = conn.execute(text(f"SELECT COUNT(*) FROM nonconformities{n_where}"), n_params).scalar_one()
    return {
        "audits": {
            "total": int(total_audits or 0),
            "by_status": {row[0]: int(row[1]) for row in audits_by_status},
        },
        "nonconformities": {
            "total": int(total_ncs or 0),
            "by_status": {row[0]: int(row[1]) for row in ncs_by_status},
            "by_severity": {row[0]: int(row[1]) for row in ncs_by_severity},
            "overdue": int(overdue or 0),
        },
        "assessments": {
            "by_status": {row[0]: int(row[1]) for row in assessments_by_status},
        },
        "generated_at": datetime.utcnow().isoformat(),
    }


@app.get("/kpi/overdue_nonconformities")
def kpi_overdue_nonconformities(days: int = 0):
    """List open nonconformities overdue by at least `days`."""
    if days < 0:
        raise HTTPException(status_code=400, detail="days must be >= 0")
    today = datetime.utcnow().date()
    cutoff = today if days == 0 else (today - timedelta(days=days))
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                "SELECT id, description, severity, status, due_date FROM nonconformities WHERE status != 'Closed' AND due_date IS NOT NULL AND due_date < :cutoff ORDER BY due_date ASC"
            ),
            {"cutoff": cutoff.isoformat()},
        ).mappings().all()
    items = []
    for r in rows:
        dd = r.get("due_date")
        try:
            d = date.fromisoformat(dd) if dd else None
        except Exception:
            d = None
        days_overdue = None
        if d is not None:
            delta = today - d
            if delta.days >= days:
                days_overdue = delta.days
            else:
                continue
        items.append({
            "id": r["id"],
            "description": r["description"],
            "severity": r["severity"],
            "status": r["status"],
            "due_date": dd,
            "days_overdue": days_overdue,
        })
    return {"count": len(items), "items": items}


@app.get("/kpi/nc_trends")
def kpi_nc_trends(from_date: Optional[date] = None, to_date: Optional[date] = None, severity: Optional[SeverityEnum] = None):
    filters_open = []
    params_open: dict = {}
    if from_date:
        filters_open.append("created_at >= :from_date")
        params_open["from_date"] = from_date.isoformat()
    if to_date:
        filters_open.append("created_at <= :to_date")
        params_open["to_date"] = to_date.isoformat()
    if severity:
        filters_open.append("severity = :severity")
        params_open["severity"] = severity.value if isinstance(severity, SeverityEnum) else severity
    where_open = (" WHERE " + " AND ".join(filters_open)) if filters_open else ""

    filters_closed = []
    params_closed: dict = {}
    if from_date:
        filters_closed.append("closed_date >= :from_date")
        params_closed["from_date"] = from_date.isoformat()
    if to_date:
        filters_closed.append("closed_date <= :to_date")
        params_closed["to_date"] = to_date.isoformat()
    if severity:
        filters_closed.append("severity = :severity")
        params_closed["severity"] = severity.value if isinstance(severity, SeverityEnum) else severity
    filters_closed.append("closed_date IS NOT NULL")
    where_closed = (" WHERE " + " AND ".join(filters_closed)) if filters_closed else " WHERE closed_date IS NOT NULL"

    with engine.connect() as conn:
        opened = conn.execute(text(f"SELECT substr(created_at,1,7) as ym, COUNT(*) as c FROM nonconformities{where_open} GROUP BY ym ORDER BY ym"), params_open).all()
        closed = conn.execute(text(f"SELECT substr(closed_date,1,7) as ym, COUNT(*) as c FROM nonconformities{where_closed} GROUP BY ym ORDER BY ym"), params_closed).all()
    opened_map = {row[0]: int(row[1]) for row in opened if row[0]}
    closed_map = {row[0]: int(row[1]) for row in closed if row[0]}
    months = sorted(set(opened_map.keys()) | set(closed_map.keys()))
    return {"months": months, "opened": opened_map, "closed": closed_map}


@app.get("/clauses", response_model=List[Clause])
def list_clauses(q: Optional[str] = None, limit: int = 50, offset: int = 0, response: Response = None):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT clause_id, title, summary FROM clauses"
    params = {}
    if q:
        base += " WHERE clause_id LIKE :q OR title LIKE :q OR summary LIKE :q"
        params["q"] = f"%{q}%"
    base += " ORDER BY clause_id LIMIT :limit OFFSET :offset"
    params.update({"limit": limit, "offset": offset})
    with engine.connect() as conn:
        total = conn.execute(text(
            ("SELECT COUNT(*) FROM clauses WHERE clause_id LIKE :q OR title LIKE :q OR summary LIKE :q") if q else ("SELECT COUNT(*) FROM clauses")
        ), params if q else {}).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    if response is not None:
        response.headers["X-Total-Count"] = str(int(total or 0))
    return [Clause(**dict(r)) for r in rows]


# --- v1 envelope endpoints (non-breaking; legacy endpoints remain) ---
router_v1 = APIRouter(prefix="/v1")


@router_v1.get("/assessments", response_model=Envelope[Assessment], tags=["Assessments"], responses={400:{"model":ErrorResponse},403:{"model":ErrorResponse}})
def v1_list_assessments(
    clause_id: Optional[str] = None,
    status: Optional[StatusEnum] = None,
    limit: int = 50,
    offset: int = 0,
    request: Request = None,
):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT * FROM assessments"
    filters = []
    params = {}
    if clause_id:
        filters.append("clause_id = :cid")
        params["cid"] = clause_id
    if status:
        filters.append("status = :status")
        params["status"] = status.value if isinstance(status, StatusEnum) else status
    org_id = getattr(request.state, "org_id", None) if request else None
    if org_id:
        filters.append("org_id = :org_id")
        params["org_id"] = org_id
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    params.update({"limit": limit, "offset": offset})
    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM assessments" + (" WHERE " + " AND ".join(filters) if filters else "")), params).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    return Envelope[Assessment](items=[_row_to_assessment(r) for r in rows], total=int(total or 0), limit=limit, offset=offset)


@router_v1.get("/audits", response_model=Envelope[Audit], tags=["Audits"], responses={400:{"model":ErrorResponse}})
def v1_list_audits(status: Optional[AuditStatus] = None, limit: int = 50, offset: int = 0, request: Request = None):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT * FROM audits"
    params = {}
    if status:
        base += " WHERE status = :status"
        params["status"] = status.value if isinstance(status, AuditStatus) else status
    org_id = getattr(request.state, "org_id", None) if request else None
    if org_id:
        base += (" WHERE " if " WHERE " not in base else " AND ") + " org_id = :org_id"
        params["org_id"] = org_id
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    params.update({"limit": limit, "offset": offset})
    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM audits" + ((" WHERE status = :status" ) if status else "") + (" WHERE org_id = :org_id" if (org_id and not status) else (" AND org_id = :org_id" if org_id else ""))), params).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    return Envelope[Audit](items=[_row_to_audit(r) for r in rows], total=int(total or 0), limit=limit, offset=offset)


@router_v1.get("/nonconformities", response_model=Envelope[Nonconformity], tags=["Nonconformities"], responses={400:{"model":ErrorResponse}})
def v1_list_nonconformities(
    status: Optional[NCStatusEnum] = None,
    severity: Optional[SeverityEnum] = None,
    clause_id: Optional[str] = None,
    audit_id: Optional[int] = None,
    limit: int = 50,
    offset: int = 0,
    request: Request = None,
):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT * FROM nonconformities"
    filters = []
    params = {}
    if status:
        filters.append("status = :status"); params["status"] = status.value if isinstance(status, NCStatusEnum) else status
    if severity:
        filters.append("severity = :severity"); params["severity"] = severity.value if isinstance(severity, SeverityEnum) else severity
    if clause_id:
        filters.append("clause_id = :cid"); params["cid"] = clause_id
    if audit_id is not None:
        filters.append("audit_id = :aid"); params["aid"] = audit_id
    org_id = getattr(request.state, "org_id", None) if request else None
    if org_id:
        filters.append("org_id = :org_id"); params["org_id"] = org_id
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    params.update({"limit": limit, "offset": offset})
    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM nonconformities" + (" WHERE " + " AND ".join(filters) if filters else "")), params).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    return Envelope[Nonconformity](items=[_row_to_nc(r) for r in rows], total=int(total or 0), limit=limit, offset=offset)


@router_v1.get("/management-reviews", response_model=Envelope[ManagementReview], tags=["Management Reviews"], responses={400:{"model":ErrorResponse}})
def v1_list_management_reviews(limit: int = 50, offset: int = 0, request: Request = None):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    order = "ORDER BY meeting_date DESC NULLS LAST, id DESC" if not DEFAULT_DB_URL.startswith("sqlite") else "ORDER BY meeting_date DESC, id DESC"
    base = f"SELECT * FROM management_reviews"
    params = {"limit": limit, "offset": offset}
    org_id = getattr(request.state, "org_id", None) if request else None
    if org_id:
        base += " WHERE org_id = :org_id"; params["org_id"] = org_id
    base += f" {order} LIMIT :limit OFFSET :offset"
    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM management_reviews" + (" WHERE org_id = :org_id" if org_id else "")), params if org_id else {}).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    return Envelope[ManagementReview](items=[_row_to_mr(r) for r in rows], total=int(total or 0), limit=limit, offset=offset)


@router_v1.get("/clauses", response_model=Envelope[Clause], tags=["Clauses"], responses={400:{"model":ErrorResponse}})
def v1_list_clauses(q: Optional[str] = None, limit: int = 50, offset: int = 0):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT clause_id, title, summary FROM clauses"
    params = {}
    if q:
        base += " WHERE clause_id LIKE :q OR title LIKE :q OR summary LIKE :q"; params["q"] = f"%{q}%"
    base += " ORDER BY clause_id LIMIT :limit OFFSET :offset"; params.update({"limit": limit, "offset": offset})
    with engine.connect() as conn:
        total = conn.execute(text(("SELECT COUNT(*) FROM clauses WHERE clause_id LIKE :q OR title LIKE :q OR summary LIKE :q") if q else ("SELECT COUNT(*) FROM clauses")), params if q else {}).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    return Envelope[Clause](items=[Clause(**dict(r)) for r in rows], total=int(total or 0), limit=limit, offset=offset)


@router_v1.post("/orgs", response_model=Organization, tags=["Organizations"], responses={400:{"model":ErrorResponse}}, summary="Create an organization")
def v1_create_org(payload: OrganizationCreate, _auth=Depends(role_required("admin"))):
    now = datetime.utcnow().isoformat()
    if not payload.id or len(payload.id) > 64:
        raise HTTPException(status_code=400, detail="invalid id")
    with engine.begin() as conn:
        exists = conn.execute(text("SELECT 1 FROM organizations WHERE id = :id"), {"id": payload.id}).first()
        if exists:
            # idempotent create returns existing
            row = conn.execute(text("SELECT * FROM organizations WHERE id = :id"), {"id": payload.id}).mappings().first()
            return Organization(**dict(row))
        conn.execute(text("INSERT INTO organizations (id, name, created_at) VALUES (:id, :name, :ts)"), {"id": payload.id, "name": payload.name, "ts": now})
        row = conn.execute(text("SELECT * FROM organizations WHERE id = :id"), {"id": payload.id}).mappings().first()
    return Organization(**dict(row))


@router_v1.get("/orgs", response_model=Envelope[Organization], tags=["Organizations"], summary="List organizations")
def v1_list_orgs(limit: int = 50, offset: int = 0, _auth=Depends(role_required("admin"))):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM organizations")).scalar_one()
        rows = conn.execute(text("SELECT * FROM organizations ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"), {"limit": limit, "offset": offset}).mappings().all()
    return Envelope[Organization](items=[Organization(**dict(r)) for r in rows], total=int(total or 0), limit=limit, offset=offset)


@router_v1.get("/orgs/{org_id}", response_model=Organization, tags=["Organizations"], responses={404:{"model":ErrorResponse}}, summary="Get organization")
def v1_get_org(org_id: str, _auth=Depends(role_required("admin"))):
    with engine.connect() as conn:
        row = conn.execute(text("SELECT * FROM organizations WHERE id = :id"), {"id": org_id}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Organization not found")
    return Organization(**dict(row))


@router_v1.patch("/orgs/{org_id}", response_model=Organization, tags=["Organizations"], responses={400:{"model":ErrorResponse},404:{"model":ErrorResponse}}, summary="Update organization")
def v1_update_org(org_id: str, payload: OrganizationUpdate, _auth=Depends(role_required("admin"))):
    updates = {}
    if payload.name is not None:
        updates["name"] = payload.name
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["id"] = org_id
    set_clause = ", ".join([f"{k} = :{k}" for k in updates.keys() if k != "id"])
    with engine.begin() as conn:
        res = conn.execute(text(f"UPDATE organizations SET {set_clause} WHERE id = :id"), updates)
        if res.rowcount == 0:
            raise HTTPException(status_code=404, detail="Organization not found")
        row = conn.execute(text("SELECT * FROM organizations WHERE id = :id"), {"id": org_id}).mappings().first()
    return Organization(**dict(row))


app.include_router(router_v1)


# --- Setup (admin) ---
class SetupRequest(BaseModel):
    run_migrations: Optional[bool] = True
    reseed_clauses: Optional[bool] = False
    verify_object_store: Optional[bool] = True


@app.post("/setup", tags=["Setup"], responses={400:{"model":ErrorResponse},401:{"model":ErrorResponse},403:{"model":ErrorResponse}})
def setup(request: Request, payload: Optional[SetupRequest] = None, _auth=Depends(role_required("admin"))):
    actions = []
    payload = payload or SetupRequest()
    # Optionally run Alembic migrations
    if payload.run_migrations:
        try:
            env = os.environ.copy()
            # Ensure DATABASE_URL propagates to alembic
            subprocess.run([
                "alembic", "-c", str(ROOT_DIR / "backend" / "alembic.ini"), "upgrade", "head"
            ], check=True, cwd=str(ROOT_DIR), env=env)
            actions.append("migrated")
        except Exception as e:
            actions.append(f"migrate_failed:{e}")
    # Ensure org exists if DEFAULT_ORG_ID set
    default_org = os.getenv("DEFAULT_ORG_ID")
    if default_org:
        with engine.begin() as conn:
            row = conn.execute(text("SELECT 1 FROM organizations WHERE id = :id"), {"id": default_org}).first()
            if not row:
                conn.execute(text("INSERT INTO organizations (id, name, created_at) VALUES (:id, :name, :ts)"), {
                    "id": default_org,
                    "name": default_org,
                    "ts": datetime.utcnow().isoformat(),
                })
                actions.append(f"created_org:{default_org}")
    # Seed clauses
    with engine.begin() as conn:
        cnt = conn.execute(text("SELECT COUNT(1) FROM clauses")).scalar_one()
        if payload.reseed_clauses:
            conn.execute(text("DELETE FROM clauses"))
            cnt = 0
    if not cnt:
        seed_clauses_if_empty()
        actions.append("seeded_clauses")
    # Verify object store
    if payload.verify_object_store:
        try:
            client, bucket = _get_s3_client()
            if client and bucket:
                client.head_bucket(Bucket=bucket)
                actions.append("object_store_ok")
            else:
                actions.append("object_store_not_configured")
        except Exception as e:
            actions.append(f"object_store_failed:{e}")
    # Return summary
    return {"ok": True, "actions": actions}


@app.get("/clauses/{clause_id}", response_model=Clause)
def get_clause(clause_id: str):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT clause_id, title, summary FROM clauses WHERE clause_id = :cid"),
            {"cid": clause_id},
        ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Clause not found")
    return Clause(**dict(row))


@app.post(
    "/assessments",
    response_model=Assessment,
    status_code=201,
    tags=["Assessments"],
    responses={400:{"model":ErrorResponse},401:{"model":ErrorResponse},403:{"model":ErrorResponse}},
    openapi_extra={
        "x-codeSamples": [
            {
                "lang": "cURL",
                "label": "curl",
                "source": "curl -X POST '{{baseUrl}}/assessments' -H 'Content-Type: application/json' -H 'Authorization: Bearer {{token}}' -d '{\\n  \"clause_id\": \"4.1\",\\n  \"status\": \"Compliant\"\\n}'"
            },
            {
                "lang": "Python",
                "label": "requests",
                "source": "import requests\nrequests.post(f'{baseUrl}/assessments', headers={'Authorization': f'Bearer {token}'}, json={'clause_id':'4.1','status':'Compliant'}).json()"
            }
        ]
    }
)
def create_assessment(payload: AssessmentCreate, request: Request, _auth=Depends(role_required("editor"))):
    # Ensure clause exists
    with engine.connect() as conn:
        exists = conn.execute(
            text("SELECT 1 FROM clauses WHERE clause_id = :cid"), {"cid": payload.clause_id}
        ).first()
        if not exists:
            raise HTTPException(status_code=400, detail="Unknown clause_id")
    now = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        res = conn.execute(
            text(
                """
                INSERT INTO assessments (clause_id, status, evidence, owner, due_date, created_at, updated_at, org_id, created_by, updated_by, request_id)
                VALUES (:clause_id, :status, :evidence, :owner, :due_date, :created_at, :updated_at, :org_id, :created_by, :updated_by, :request_id)
                RETURNING id
                """
                if not DEFAULT_DB_URL.startswith("sqlite")
                else """
                INSERT INTO assessments (clause_id, status, evidence, owner, due_date, created_at, updated_at, org_id, created_by, updated_by, request_id)
                VALUES (:clause_id, :status, :evidence, :owner, :due_date, :created_at, :updated_at, :org_id, :created_by, :updated_by, :request_id)
                """
            ),
            {
                "clause_id": payload.clause_id,
                "status": payload.status.value if isinstance(payload.status, StatusEnum) else payload.status,
                "evidence": payload.evidence,
                "owner": payload.owner,
                "due_date": payload.due_date.isoformat() if payload.due_date else None,
                "created_at": now,
                "updated_at": now,
                "org_id": getattr(request.state, "org_id", None),
                "created_by": (_auth or {}).get("sub"),
                "updated_by": (_auth or {}).get("sub"),
                "request_id": getattr(request.state, "request_id", None),
            },
        )
        if DEFAULT_DB_URL.startswith("sqlite"):
            new_id = conn.execute(text("SELECT last_insert_rowid()"))
            new_id = list(new_id)[0][0]
        else:
            new_id = res.scalar_one()
        row = conn.execute(text("SELECT * FROM assessments WHERE id = :id"), {"id": new_id}).mappings().first()
    return Assessment(**dict(row))


def _row_to_assessment(row) -> Assessment:
    d = dict(row)
    if d.get("due_date"):
        try:
            d["due_date"] = date.fromisoformat(d["due_date"])  # type: ignore
        except Exception:
            d["due_date"] = None
    return Assessment(**d)


@app.get("/assessments", response_model=List[Assessment])
def list_assessments(
    clause_id: Optional[str] = None,
    status: Optional[StatusEnum] = None,
    limit: int = 50,
    offset: int = 0,
    response: Response = None,
    request: Request = None,
):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT * FROM assessments"
    filters = []
    params = {}
    if clause_id:
        filters.append("clause_id = :cid")
        params["cid"] = clause_id
    if status:
        filters.append("status = :status")
        params["status"] = status.value if isinstance(status, StatusEnum) else status
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    params.update({"limit": limit, "offset": offset})
    with engine.connect() as conn:
        count_sql = "SELECT COUNT(*) FROM assessments" + (" WHERE " + " AND ".join(filters) if filters else "")
        total = conn.execute(text(count_sql), params).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    if response is not None:
        response.headers["X-Total-Count"] = str(int(total or 0))
    return [_row_to_assessment(r) for r in rows]


@app.get("/assessments/{assessment_id}", response_model=Assessment)
def get_assessment(assessment_id: int, request: Request = None):
    with engine.connect() as conn:
        params = {"id": assessment_id}
        where = "id = :id"
        org_id = getattr(request.state, "org_id", None) if request else None
        if org_id:
            where += " AND org_id = :org_id"
            params["org_id"] = org_id
        row = conn.execute(
            text(f"SELECT * FROM assessments WHERE {where}"), params
        ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Assessment not found")
    return _row_to_assessment(row)


# --- Evidence attachments ---
def _ensure_evidence_dir():
    EVIDENCE_DIR.mkdir(parents=True, exist_ok=True)


# --- Basic in-memory rate limiter for sensitive routes ---
_RATE_STATE = {}

def rate_limit(key: str, limit: int = 30, window_sec: int = 60):
    # Prefer Redis token bucket if REDIS_URL is set, fallback to in-memory
    redis_url = os.getenv("REDIS_URL")
    if redis_url:
        try:
            import redis  # type: ignore
            r = redis.Redis.from_url(redis_url)
            now = int(time.time())
            window = now // window_sec
            rk = f"ratelimit:{key}:{window}"
            val = r.incr(rk)
            if val == 1:
                r.expire(rk, window_sec)
            if val > limit:
                raise HTTPException(status_code=429, detail="Rate limit exceeded")
            return
        except Exception:
            pass
    import time as _t
    now = int(_t.time())
    window = now // window_sec
    k = (key, window)
    count = _RATE_STATE.get(k, 0) + 1
    _RATE_STATE[k] = count
    prev = (key, window - 1)
    _RATE_STATE.pop(prev, None)
    if count > limit:
        raise HTTPException(status_code=429, detail="Rate limit exceeded")


@app.post(
    "/assessments/{assessment_id}/attachments",
    response_model=Attachment,
    status_code=201,
    tags=["Attachments"],
    responses={400:{"model":ErrorResponse},401:{"model":ErrorResponse},403:{"model":ErrorResponse},413:{"model":ErrorResponse},429:{"model":ErrorResponse}},
    openapi_extra={
        "x-codeSamples": [
            {"lang":"cURL","label":"curl (multipart)","source":"curl -X POST '{{baseUrl}}/assessments/1/attachments' -H 'Authorization: Bearer {{token}}' -F 'file=@evidence.pdf;type=application/pdf'"}
        ]
    }
)
def upload_attachment(assessment_id: int, request: Request, file: UploadFile = File(...), _auth=Depends(role_required("editor"))):
    # Rate limit and size check
    rate_limit(f"upload:{request.client.host if request.client else 'local'}")
    max_mb = int(os.getenv("MAX_UPLOAD_MB", "25"))
    cl = request.headers.get("content-length")
    if cl and int(cl) > max_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Payload too large")
    with engine.connect() as conn:
        exists = conn.execute(text("SELECT 1 FROM assessments WHERE id = :id"), {"id": assessment_id}).first()
        if not exists:
            raise HTTPException(status_code=404, detail="Assessment not found")
    _ensure_evidence_dir()
    subdir = EVIDENCE_DIR / f"assessment_{assessment_id}"
    subdir.mkdir(parents=True, exist_ok=True)
    ext = Path(file.filename).suffix if file.filename else ""
    uid = uuid.uuid4().hex
    stored_name = f"{uid}{ext}"
    dest_path = subdir / stored_name
    with dest_path.open("wb") as out:
        shutil.copyfileobj(file.file, out)
    size = dest_path.stat().st_size
    now = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        res = conn.execute(
            text(
                """
                INSERT INTO attachments (assessment_id, filename, content_type, size, stored_path, created_at, org_id, created_by, request_id)
                VALUES (:aid, :filename, :content_type, :size, :stored_path, :created_at, :org_id, :created_by, :request_id)
                RETURNING id
                """
                if not DEFAULT_DB_URL.startswith("sqlite")
                else """
                INSERT INTO attachments (assessment_id, filename, content_type, size, stored_path, created_at, org_id, created_by, request_id)
                VALUES (:aid, :filename, :content_type, :size, :stored_path, :created_at, :org_id, :created_by, :request_id)
                """
            ),
            {
                "aid": assessment_id,
                "filename": file.filename or stored_name,
                "content_type": file.content_type,
                "size": int(size),
                "stored_path": str(dest_path),
                "created_at": now,
                "org_id": getattr(request.state, "org_id", None),
                "created_by": (_auth or {}).get("sub"),
                "request_id": getattr(request.state, "request_id", None),
            },
        )
        if DEFAULT_DB_URL.startswith("sqlite"):
            att_id = conn.execute(text("SELECT last_insert_rowid()")).scalar_one()
        else:
            att_id = res.scalar_one()
        row = conn.execute(text("SELECT id, assessment_id, filename, content_type, size, created_at FROM attachments WHERE id = :id"), {"id": att_id}).mappings().first()
    return Attachment(**dict(row))


@app.get("/assessments/{assessment_id}/attachments", response_model=List[Attachment])
def list_attachments(assessment_id: int, response: Response = None, request: Request = None):
    with engine.connect() as conn:
        exists = conn.execute(text("SELECT 1 FROM assessments WHERE id = :id"), {"id": assessment_id}).first()
        if not exists:
            raise HTTPException(status_code=404, detail="Assessment not found")
        params = {"id": assessment_id}
        where = "assessment_id = :id"
        org_id = getattr(request.state, "org_id", None) if request else None
        if org_id:
            where += " AND org_id = :org_id"
            params["org_id"] = org_id
        rows = conn.execute(
            text(f"SELECT id, assessment_id, filename, content_type, size, created_at FROM attachments WHERE {where} ORDER BY created_at DESC, id DESC"),
            params,
        ).mappings().all()
        total = conn.execute(text(f"SELECT COUNT(*) FROM attachments WHERE {where}"), params).scalar_one()
    if response is not None:
        response.headers["X-Total-Count"] = str(int(total or 0))
    return [Attachment(**dict(r)) for r in rows]


@app.get("/attachments/{attachment_id}/download")
def download_attachment(attachment_id: int):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT filename, content_type, stored_path FROM attachments WHERE id = :id"),
            {"id": attachment_id},
        ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")
    path = Path(row["stored_path"]).resolve()
    if not path.exists():
        raise HTTPException(status_code=410, detail="File no longer available")
    return FileResponse(str(path), media_type=row.get("content_type") or "application/octet-stream", filename=row.get("filename") or path.name)


@app.delete("/attachments/{attachment_id}", status_code=204)
def delete_attachment(attachment_id: int, request: Request, _auth=Depends(role_required("admin"))):
    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT stored_path, org_id FROM attachments WHERE id = :id"), {"id": attachment_id}
        ).mappings().first()
        if not row:
            raise HTTPException(status_code=404, detail="Attachment not found")
        # org enforcement (if present)
        req_org = getattr(request, 'state', None) and getattr(request.state, 'org_id', None)
        if row.get("org_id") and req_org and row.get("org_id") != req_org:
            raise HTTPException(status_code=403, detail="Forbidden: wrong org")
        conn.execute(text("DELETE FROM attachments WHERE id = :id"), {"id": attachment_id})
    try:
        Path(row["stored_path"]).unlink(missing_ok=True)
    except Exception:
        pass
    return {}


@app.post(
    "/attachments/presign_upload",
    response_model=PresignResponse,
    tags=["Attachments"],
    responses={400:{"model":ErrorResponse},401:{"model":ErrorResponse},403:{"model":ErrorResponse},429:{"model":ErrorResponse}},
    openapi_extra={
        "x-codeSamples": [
            {"lang":"cURL","label":"curl","source":"curl -X POST '{{baseUrl}}/attachments/presign_upload' -H 'Authorization: Bearer {{token}}' -H 'Content-Type: application/json' -d '{\\n  \\\"assessment_id\\\": 1,\\n  \\\"filename\\\": \\\"evidence.pdf\\\",\\n  \\\"content_type\\\": \\\"application/pdf\\\",\\n  \\\"size\\\": 1048576\\n}'"}
        ]
    }
)
def presign_upload(payload: PresignRequest, request: Request, _auth=Depends(role_required("editor"))):
    # Rate limit
    rate_limit(f"presign:{request.client.host if request.client else 'local'}")
    client, bucket = _get_s3_client()
    if not client:
        raise HTTPException(status_code=400, detail="Object store not configured")
    # Basic sanitization
    fname = re.sub(r"[^A-Za-z0-9_.-]", "_", payload.filename)
    org_id = getattr(request.state, "org_id", "public") or "public"
    key = f"{org_id}/assessments/{payload.assessment_id}/{uuid.uuid4().hex}_{fname}"
    # MIME allow-list and size limits
    allowed_env = os.getenv("ALLOWED_UPLOAD_MEDIA")
    allowed = {m.strip() for m in allowed_env.split(",") if m.strip()} if allowed_env else {
        "application/pdf", "image/png", "image/jpeg",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "text/csv",
    }
    if payload.content_type and payload.content_type not in allowed:
        raise HTTPException(status_code=400, detail=f"content_type not allowed: {payload.content_type}")
    max_mb = int(os.getenv("MAX_UPLOAD_MB", "25"))
    if payload.size and payload.size > max_mb * 1024 * 1024:
        raise HTTPException(status_code=400, detail="file too large")
    params = {"Bucket": bucket, "Key": key}
    # Server-side encryption (optional)
    sse = os.getenv("OBJECT_STORE_SSE")  # AES256 or aws:kms
    kms_key = os.getenv("OBJECT_STORE_KMS_KEY")
    if sse:
        params["ServerSideEncryption"] = sse
        if sse == "aws:kms" and kms_key:
            params["SSEKMSKeyId"] = kms_key
    if payload.content_type:
        params["ContentType"] = payload.content_type
    ttl = int(os.getenv("PRESIGN_TTL_SECONDS", "900"))
    url = client.generate_presigned_url("put_object", Params=params, ExpiresIn=ttl)
    headers = {}
    if payload.content_type:
        headers["Content-Type"] = payload.content_type
    if sse:
        # Clients must include these headers when uploading with presigned URL
        headers["x-amz-server-side-encryption"] = sse
        if sse == "aws:kms" and kms_key:
            headers["x-amz-server-side-encryption-aws-kms-key-id"] = kms_key
    return PresignResponse(upload_url=url, object_key=key, headers=headers)


@app.post(
    "/attachments/complete",
    response_model=Attachment,
    status_code=201,
    tags=["Attachments"],
    responses={400:{"model":ErrorResponse},401:{"model":ErrorResponse},403:{"model":ErrorResponse}},
    openapi_extra={
        "x-codeSamples": [
            {"lang":"cURL","label":"curl","source":"curl -X POST '{{baseUrl}}/attachments/complete' -H 'Authorization: Bearer {{token}}' -H 'Content-Type: application/json' -d '{\\n  \\\"assessment_id\\\": 1,\\n  \\\"object_key\\\": \\\"orgA/assessments/1/abc123_evidence.pdf\\\",\\n  \\\"filename\\\": \\\"evidence.pdf\\\",\\n  \\\"content_type\\\": \\\"application/pdf\\\",\\n  \\\"size\\\": 1048576\\n}'"}
        ]
    }
)
def complete_attachment(payload: AttachmentComplete, request: Request, _auth=Depends(role_required("editor"))):
    # Record the attachment metadata pointing to object storage
    now = datetime.utcnow().isoformat()
    stored_path = f"s3://{os.getenv('OBJECT_STORE_BUCKET','')}/{payload.object_key}"
    with engine.begin() as conn:
        res = conn.execute(
            text(
                """
                INSERT INTO attachments (assessment_id, filename, content_type, size, stored_path, created_at, org_id, created_by, request_id, sha256, retention_hold, retention_until, disposition)
                VALUES (:aid, :filename, :content_type, :size, :stored_path, :created_at, :org_id, :created_by, :request_id, :sha256, :retention_hold, :retention_until, :disposition)
                RETURNING id
                """
                if not DEFAULT_DB_URL.startswith("sqlite")
                else """
                INSERT INTO attachments (assessment_id, filename, content_type, size, stored_path, created_at, org_id, created_by, request_id, sha256, retention_hold, retention_until, disposition)
                VALUES (:aid, :filename, :content_type, :size, :stored_path, :created_at, :org_id, :created_by, :request_id, :sha256, :retention_hold, :retention_until, :disposition)
                """
            ),
            {
                "aid": payload.assessment_id,
                "filename": payload.filename,
                "content_type": payload.content_type,
                "size": payload.size,
                "stored_path": stored_path,
                "created_at": now,
                "org_id": getattr(request.state, "org_id", None),
                "created_by": (_auth or {}).get("sub"),
                "request_id": getattr(request.state, "request_id", None),
                "sha256": payload.sha256,
                "retention_hold": str(payload.retention_hold) if payload.retention_hold is not None else None,
                "retention_until": payload.retention_until.isoformat() if payload.retention_until else None,
                "disposition": payload.disposition,
            },
        )
        if DEFAULT_DB_URL.startswith("sqlite"):
            att_id = conn.execute(text("SELECT last_insert_rowid()")).scalar_one()
        else:
            att_id = res.scalar_one()
        row = conn.execute(text("SELECT id, assessment_id, filename, content_type, size, created_at, stored_path FROM attachments WHERE id = :id"), {"id": att_id}).mappings().first()
    # Optional AV scan hook
    try:
        scan_cmd = os.getenv("AV_SCAN_CMD")  # e.g., "clamscan {path}"
        if scan_cmd and row and row.get("stored_path") and row.get("stored_path").startswith("s3://"):
            # Skipping direct S3 download; AV hook best suited for local uploads or async workers
            pass
        elif scan_cmd and row and row.get("stored_path") and not row.get("stored_path").startswith("s3://"):
            spath = row.get("stored_path")
            cmd = scan_cmd.format(path=spath)
            import subprocess as _sub
            rc = _sub.call(cmd, shell=True)
            if rc != 0:
                raise HTTPException(status_code=400, detail="AV scan failed")
    except HTTPException:
        raise
    except Exception:
        # Do not fail uploads on scan errors unless explicitly configured
        pass
    return Attachment(**dict(row))


@app.get("/attachments/{attachment_id}/download_url")
def get_attachment_download_url(attachment_id: int):
    client, bucket = _get_s3_client()
    if not client:
        raise HTTPException(status_code=400, detail="Object store not configured")
    with engine.connect() as conn:
        row = conn.execute(text("SELECT filename, stored_path, content_type FROM attachments WHERE id = :id"), {"id": attachment_id}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Attachment not found")
    stored = row["stored_path"] or ""
    if not stored.startswith("s3://"):
        raise HTTPException(status_code=400, detail="Attachment is file-based; use /attachments/{id}/download")
    # parse key
    prefix = f"s3://{bucket}/"
    key = stored[len("s3://"):]
    # When bucket is included in stored_path, remove it
    if key.startswith(f"{bucket}/"):
        key = key[len(f"{bucket}/"):]
    url = client.generate_presigned_url(
        "get_object",
        Params={"Bucket": bucket, "Key": key},
        ExpiresIn=900,
    )
    return {"url": url}


@app.patch("/assessments/{assessment_id}", response_model=Assessment)
def update_assessment(assessment_id: int, payload: AssessmentUpdate, request: Request, _auth=Depends(role_required("editor"))):
    updates = {}
    if payload.status is not None:
        updates["status"] = payload.status.value if hasattr(payload.status, "value") else payload.status
    if payload.evidence is not None:
        updates["evidence"] = payload.evidence
    if payload.owner is not None:
        updates["owner"] = payload.owner
    if payload.due_date is not None:
        updates["due_date"] = payload.due_date.isoformat() if payload.due_date else None
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.utcnow().isoformat()
    updates["updated_by"] = (_auth or {}).get("sub")
    updates["request_id"] = getattr(request.state, "request_id", None)
    updates["id"] = assessment_id
    set_clause = ", ".join([f"{k} = :{k}" for k in updates.keys() if k != "id"])
    with engine.begin() as conn:
        # org enforcement
        org_id = getattr(request.state, "org_id", None)
        if org_id:
            cur = conn.execute(text("SELECT org_id FROM assessments WHERE id = :id"), {"id": assessment_id}).mappings().first()
            if cur and cur.get("org_id") and cur.get("org_id") != org_id:
                raise HTTPException(status_code=403, detail="Forbidden: wrong org")
        res = conn.execute(text(f"UPDATE assessments SET {set_clause} WHERE id = :id"), updates)
        if res.rowcount == 0:
            raise HTTPException(status_code=404, detail="Assessment not found")
        row = conn.execute(
            text("SELECT * FROM assessments WHERE id = :id"), {"id": assessment_id}
        ).mappings().first()
    return _row_to_assessment(row)


# --- Audits ---
def _row_to_audit(row) -> Audit:
    d = dict(row)
    for k in ("scheduled_date", "completed_date"):
        if d.get(k):
            try:
                d[k] = date.fromisoformat(d[k])  # type: ignore
            except Exception:
                d[k] = None
    return Audit(**d)


@app.post("/audits", response_model=Audit, status_code=201)
def create_audit(payload: AuditCreate, request: Request, _auth=Depends(role_required("editor"))):
    now = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        res = conn.execute(
            text(
                """
                INSERT INTO audits (title, description, status, scheduled_date, created_at, updated_at, org_id, created_by, updated_by, request_id)
                VALUES (:title, :description, :status, :scheduled_date, :created_at, :updated_at, :org_id, :created_by, :updated_by, :request_id)
                RETURNING id
                """
                if not DEFAULT_DB_URL.startswith("sqlite")
                else """
                INSERT INTO audits (title, description, status, scheduled_date, created_at, updated_at, org_id, created_by, updated_by, request_id)
                VALUES (:title, :description, :status, :scheduled_date, :created_at, :updated_at, :org_id, :created_by, :updated_by, :request_id)
                """
            ),
            {
                "title": payload.title,
                "description": payload.description,
                "status": payload.status.value if isinstance(payload.status, AuditStatus) else payload.status,
                "scheduled_date": payload.scheduled_date.isoformat() if payload.scheduled_date else None,
                "created_at": now,
                "updated_at": now,
                "org_id": getattr(request.state, "org_id", None),
                "created_by": (_auth or {}).get("sub"),
                "updated_by": (_auth or {}).get("sub"),
                "request_id": getattr(request.state, "request_id", None),
            },
        )
        if DEFAULT_DB_URL.startswith("sqlite"):
            new_id = conn.execute(text("SELECT last_insert_rowid()")).scalar_one()
        else:
            new_id = res.scalar_one()
        row = conn.execute(text("SELECT * FROM audits WHERE id = :id"), {"id": new_id}).mappings().first()
    return _row_to_audit(row)


@app.get("/audits", response_model=List[Audit])
def list_audits(status: Optional[AuditStatus] = None, limit: int = 50, offset: int = 0, response: Response = None, request: Request = None):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT * FROM audits"
    params = {}
    if status:
        base += " WHERE status = :status"
        params["status"] = status.value if isinstance(status, AuditStatus) else status
    # org scope
    org_id = getattr(request.state, "org_id", None) if request else None
    if org_id:
        base += (" WHERE " if " WHERE " not in base else " AND ") + " org_id = :org_id"
        params["org_id"] = org_id
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    params.update({"limit": limit, "offset": offset})
    with engine.connect() as conn:
        count_sql = "SELECT COUNT(*) FROM audits" + (" WHERE status = :status" if status else "")
        if org_id:
            count_sql += (" WHERE " if " WHERE " not in count_sql else " AND ") + " org_id = :org_id"
        total = conn.execute(text(count_sql), params).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    if response is not None:
        response.headers["X-Total-Count"] = str(int(total or 0))
    return [_row_to_audit(r) for r in rows]


@app.get("/audits/{audit_id}", response_model=Audit)
def get_audit(audit_id: int):
    with engine.connect() as conn:
        row = conn.execute(text("SELECT * FROM audits WHERE id = :id"), {"id": audit_id}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Audit not found")
    return _row_to_audit(row)


@app.patch("/audits/{audit_id}", response_model=Audit)
def update_audit(audit_id: int, payload: AuditUpdate, request: Request, _auth=Depends(role_required("editor"))):
    updates = {}
    if payload.title is not None:
        updates["title"] = payload.title
    if payload.description is not None:
        updates["description"] = payload.description
    if payload.status is not None:
        updates["status"] = payload.status.value if isinstance(payload.status, AuditStatus) else payload.status
    if payload.scheduled_date is not None:
        updates["scheduled_date"] = payload.scheduled_date.isoformat() if payload.scheduled_date else None
    if payload.completed_date is not None:
        updates["completed_date"] = payload.completed_date.isoformat() if payload.completed_date else None
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.utcnow().isoformat()
    updates["updated_by"] = (_auth or {}).get("sub")
    updates["request_id"] = getattr(request.state, "request_id", None)
    updates["id"] = audit_id
    set_clause = ", ".join([f"{k} = :{k}" for k in updates.keys() if k != "id"])
    with engine.begin() as conn:
        org_id = getattr(request.state, "org_id", None)
        if org_id:
            cur = conn.execute(text("SELECT org_id FROM audits WHERE id = :id"), {"id": audit_id}).mappings().first()
            if cur and cur.get("org_id") and cur.get("org_id") != org_id:
                raise HTTPException(status_code=403, detail="Forbidden: wrong org")
        res = conn.execute(text(f"UPDATE audits SET {set_clause} WHERE id = :id"), updates)
        if res.rowcount == 0:
            raise HTTPException(status_code=404, detail="Audit not found")
        row = conn.execute(text("SELECT * FROM audits WHERE id = :id"), {"id": audit_id}).mappings().first()
    return _row_to_audit(row)


# --- Nonconformities ---
def _row_to_nc(row) -> Nonconformity:
    d = dict(row)
    for k in ("due_date", "closed_date", "verified_on"):
        if d.get(k):
            try:
                d[k] = date.fromisoformat(d[k])  # type: ignore
            except Exception:
                d[k] = None
    return Nonconformity(**d)


@app.post("/nonconformities", response_model=Nonconformity, status_code=201)
def create_nonconformity(payload: NonconformityCreate, request: Request, _auth=Depends(role_required("editor"))):
    # Validate foreign keys (audit_id, clause_id) if provided
    with engine.connect() as conn:
        if payload.audit_id is not None:
            exists = conn.execute(text("SELECT 1 FROM audits WHERE id = :id"), {"id": payload.audit_id}).first()
            if not exists:
                raise HTTPException(status_code=400, detail="Unknown audit_id")
        if payload.clause_id is not None:
            exists = conn.execute(text("SELECT 1 FROM clauses WHERE clause_id = :cid"), {"cid": payload.clause_id}).first()
            if not exists:
                raise HTTPException(status_code=400, detail="Unknown clause_id")
    now = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        res = conn.execute(
            text(
                """
                INSERT INTO nonconformities (audit_id, clause_id, severity, status, state, description, corrective_action, containment, root_cause, preventive_action, verification_method, owner, due_date, created_at, updated_at, org_id, created_by, updated_by, request_id)
                VALUES (:audit_id, :clause_id, :severity, :status, :state, :description, :corrective_action, :containment, :root_cause, :preventive_action, :verification_method, :owner, :due_date, :created_at, :updated_at, :org_id, :created_by, :updated_by, :request_id)
                RETURNING id
                """
                if not DEFAULT_DB_URL.startswith("sqlite")
                else """
                INSERT INTO nonconformities (audit_id, clause_id, severity, status, state, description, corrective_action, containment, root_cause, preventive_action, verification_method, owner, due_date, created_at, updated_at, org_id, created_by, updated_by, request_id)
                VALUES (:audit_id, :clause_id, :severity, :status, :state, :description, :corrective_action, :containment, :root_cause, :preventive_action, :verification_method, :owner, :due_date, :created_at, :updated_at, :org_id, :created_by, :updated_by, :request_id)
                """
            ),
            {
                "audit_id": payload.audit_id,
                "clause_id": payload.clause_id,
                "severity": payload.severity.value if isinstance(payload.severity, SeverityEnum) else payload.severity,
                "status": payload.status.value if isinstance(payload.status, NCStatusEnum) else payload.status,
                "state": (payload.state.value if isinstance(payload.state, NCStateEnum) else payload.state) if payload.state else NCStateEnum.New.value,
                "description": payload.description,
                "corrective_action": payload.corrective_action,
                "containment": payload.containment,
                "root_cause": payload.root_cause,
                "preventive_action": payload.preventive_action,
                "verification_method": payload.verification_method,
                "owner": payload.owner,
                "due_date": payload.due_date.isoformat() if payload.due_date else None,
                "created_at": now,
                "updated_at": now,
                "org_id": getattr(request.state, "org_id", None),
                "created_by": (_auth or {}).get("sub"),
                "updated_by": (_auth or {}).get("sub"),
                "request_id": getattr(request.state, "request_id", None),
            },
        )
        if DEFAULT_DB_URL.startswith("sqlite"):
            new_id = conn.execute(text("SELECT last_insert_rowid()")).scalar_one()
        else:
            new_id = res.scalar_one()
        row = conn.execute(text("SELECT * FROM nonconformities WHERE id = :id"), {"id": new_id}).mappings().first()
    return _row_to_nc(row)


@app.get("/nonconformities", response_model=List[Nonconformity])
def list_nonconformities(
    status: Optional[NCStatusEnum] = None,
    severity: Optional[SeverityEnum] = None,
    clause_id: Optional[str] = None,
    audit_id: Optional[int] = None,
    limit: int = 50,
    offset: int = 0,
    response: Response = None,
    request: Request = None,
):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT * FROM nonconformities"
    filters = []
    params = {}
    if status:
        filters.append("status = :status")
        params["status"] = status.value if isinstance(status, NCStatusEnum) else status
    if severity:
        filters.append("severity = :severity")
        params["severity"] = severity.value if isinstance(severity, SeverityEnum) else severity
    if clause_id:
        filters.append("clause_id = :cid")
        params["cid"] = clause_id
    if audit_id is not None:
        filters.append("audit_id = :aid")
        params["aid"] = audit_id
    if filters:
        base += " WHERE " + " AND ".join(filters)
    org_id = getattr(request.state, "org_id", None) if request else None
    if org_id:
        filters.append("org_id = :org_id")
        params["org_id"] = org_id
    base += " ORDER BY created_at DESC, id DESC LIMIT :limit OFFSET :offset"
    params.update({"limit": limit, "offset": offset})
    with engine.connect() as conn:
        count_sql = "SELECT COUNT(*) FROM nonconformities" + (" WHERE " + " AND ".join(filters) if filters else "")
        total = conn.execute(text(count_sql), params).scalar_one()
        rows = conn.execute(text(base), params).mappings().all()
    if response is not None:
        response.headers["X-Total-Count"] = str(int(total or 0))
    return [_row_to_nc(r) for r in rows]


@app.get("/nonconformities/{nc_id}", response_model=Nonconformity)
def get_nonconformity(nc_id: int):
    with engine.connect() as conn:
        row = conn.execute(text("SELECT * FROM nonconformities WHERE id = :id"), {"id": nc_id}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Nonconformity not found")
    return _row_to_nc(row)


@app.patch("/nonconformities/{nc_id}", response_model=Nonconformity)
def update_nonconformity(nc_id: int, payload: NonconformityUpdate, request: Request, _auth=Depends(role_required("editor"))):
    updates = {}
    if payload.description is not None:
        updates["description"] = payload.description
    if payload.severity is not None:
        updates["severity"] = payload.severity.value if isinstance(payload.severity, SeverityEnum) else payload.severity
    if payload.status is not None:
        updates["status"] = payload.status.value if isinstance(payload.status, NCStatusEnum) else payload.status
    if payload.state is not None:
        # Enforce allowed transitions
        with engine.connect() as conn:
            cur = conn.execute(text("SELECT state FROM nonconformities WHERE id = :id"), {"id": nc_id}).mappings().first()
        curr = (cur or {}).get("state")
        allowed = {
            None: {NCStateEnum.New.value},
            NCStateEnum.New.value: {NCStateEnum.Analysis.value, NCStateEnum.Action.value},
            NCStateEnum.Analysis.value: {NCStateEnum.Action.value, NCStateEnum.Verification.value},
            NCStateEnum.Action.value: {NCStateEnum.Verification.value},
            NCStateEnum.Verification.value: {NCStateEnum.Closed.value, NCStateEnum.Action.value},
            NCStateEnum.Closed.value: set(),
        }
        target = payload.state.value if isinstance(payload.state, NCStateEnum) else str(payload.state)
        if curr not in allowed or target not in allowed[curr]:
            raise HTTPException(status_code=400, detail=f"Invalid state transition: {curr} -> {target}")
        updates["state"] = target
    if payload.audit_id is not None:
        # Validate audit exists
        with engine.connect() as conn:
            exists = conn.execute(text("SELECT 1 FROM audits WHERE id = :id"), {"id": payload.audit_id}).first()
            if not exists:
                raise HTTPException(status_code=400, detail="Unknown audit_id")
        updates["audit_id"] = payload.audit_id
    if payload.clause_id is not None:
        with engine.connect() as conn:
            exists = conn.execute(text("SELECT 1 FROM clauses WHERE clause_id = :cid"), {"cid": payload.clause_id}).first()
            if not exists:
                raise HTTPException(status_code=400, detail="Unknown clause_id")
        updates["clause_id"] = payload.clause_id
    if payload.corrective_action is not None:
        updates["corrective_action"] = payload.corrective_action
    if payload.containment is not None:
        updates["containment"] = payload.containment
    if payload.root_cause is not None:
        updates["root_cause"] = payload.root_cause
    if payload.preventive_action is not None:
        updates["preventive_action"] = payload.preventive_action
    if payload.verification_method is not None:
        updates["verification_method"] = payload.verification_method
    if payload.verified_by is not None:
        updates["verified_by"] = payload.verified_by
    if payload.verified_on is not None:
        updates["verified_on"] = payload.verified_on.isoformat() if payload.verified_on else None
    if payload.owner is not None:
        updates["owner"] = payload.owner
    if payload.due_date is not None:
        updates["due_date"] = payload.due_date.isoformat() if payload.due_date else None
    if payload.closed_date is not None:
        updates["closed_date"] = payload.closed_date.isoformat() if payload.closed_date else None
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.utcnow().isoformat()
    updates["updated_by"] = (_auth or {}).get("sub")
    updates["request_id"] = getattr(request.state, "request_id", None)
    updates["id"] = nc_id
    set_clause = ", ".join([f"{k} = :{k}" for k in updates.keys() if k != "id"])
    with engine.begin() as conn:
        org_id = getattr(request.state, "org_id", None)
        if org_id:
            cur = conn.execute(text("SELECT org_id FROM nonconformities WHERE id = :id"), {"id": nc_id}).mappings().first()
            if cur and cur.get("org_id") and cur.get("org_id") != org_id:
                raise HTTPException(status_code=403, detail="Forbidden: wrong org")
        res = conn.execute(text(f"UPDATE nonconformities SET {set_clause} WHERE id = :id"), updates)
        if res.rowcount == 0:
            raise HTTPException(status_code=404, detail="Nonconformity not found")
        row = conn.execute(text("SELECT * FROM nonconformities WHERE id = :id"), {"id": nc_id}).mappings().first()
    return _row_to_nc(row)


# --- Management Reviews ---
def _row_to_mr(row) -> ManagementReview:
    d = dict(row)
    for k in ("period_start", "period_end", "meeting_date"):
        if d.get(k):
            try:
                d[k] = date.fromisoformat(d[k])  # type: ignore
            except Exception:
                d[k] = None
    return ManagementReview(**d)


@app.post("/management-reviews", response_model=ManagementReview, status_code=201)
def create_management_review(payload: ManagementReviewCreate, request: Request, _auth=Depends(role_required("editor"))):
    now = datetime.utcnow().isoformat()
    with engine.begin() as conn:
        res = conn.execute(
            text(
                """
                INSERT INTO management_reviews (title, period_start, period_end, meeting_date, participants, summary, decisions, actions, created_at, updated_at, org_id, created_by, updated_by, request_id)
                VALUES (:title, :period_start, :period_end, :meeting_date, :participants, :summary, :decisions, :actions, :created_at, :updated_at, :org_id, :created_by, :updated_by, :request_id)
                RETURNING id
                """
                if not DEFAULT_DB_URL.startswith("sqlite")
                else """
                INSERT INTO management_reviews (title, period_start, period_end, meeting_date, participants, summary, decisions, actions, created_at, updated_at, org_id, created_by, updated_by, request_id)
                VALUES (:title, :period_start, :period_end, :meeting_date, :participants, :summary, :decisions, :actions, :created_at, :updated_at, :org_id, :created_by, :updated_by, :request_id)
                """
            ),
            {
                "title": payload.title,
                "period_start": payload.period_start.isoformat() if payload.period_start else None,
                "period_end": payload.period_end.isoformat() if payload.period_end else None,
                "meeting_date": payload.meeting_date.isoformat() if payload.meeting_date else None,
                "participants": payload.participants,
                "summary": payload.summary,
                "decisions": payload.decisions,
                "actions": payload.actions,
                "created_at": now,
                "updated_at": now,
                "org_id": getattr(request.state, "org_id", None),
                "created_by": (_auth or {}).get("sub"),
                "updated_by": (_auth or {}).get("sub"),
                "request_id": getattr(request.state, "request_id", None),
            },
        )
        if DEFAULT_DB_URL.startswith("sqlite"):
            new_id = conn.execute(text("SELECT last_insert_rowid()")).scalar_one()
        else:
            new_id = res.scalar_one()
        row = conn.execute(text("SELECT * FROM management_reviews WHERE id = :id"), {"id": new_id}).mappings().first()
    return _row_to_mr(row)


@app.get("/management-reviews", response_model=List[ManagementReview])
def list_management_reviews(limit: int = 50, offset: int = 0, response: Response = None):
    if limit < 1 or limit > 200:
        raise HTTPException(status_code=400, detail="limit must be between 1 and 200")
    if offset < 0:
        raise HTTPException(status_code=400, detail="offset must be >= 0")
    base = "SELECT * FROM management_reviews ORDER BY meeting_date DESC NULLS LAST, id DESC LIMIT :limit OFFSET :offset" if not DEFAULT_DB_URL.startswith("sqlite") else "SELECT * FROM management_reviews ORDER BY meeting_date DESC, id DESC LIMIT :limit OFFSET :offset"
    with engine.connect() as conn:
        rows = conn.execute(text(base), {"limit": limit, "offset": offset}).mappings().all()
        total = conn.execute(text("SELECT COUNT(*) FROM management_reviews")).scalar_one()
    if response is not None:
        response.headers["X-Total-Count"] = str(int(total or 0))
    return [_row_to_mr(r) for r in rows]


@app.get("/management-reviews/{mr_id}", response_model=ManagementReview)
def get_management_review(mr_id: int):
    with engine.connect() as conn:
        row = conn.execute(text("SELECT * FROM management_reviews WHERE id = :id"), {"id": mr_id}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Management review not found")
    return _row_to_mr(row)


@app.patch("/management-reviews/{mr_id}", response_model=ManagementReview)
def update_management_review(mr_id: int, payload: ManagementReviewUpdate, request: Request, _auth=Depends(role_required("editor"))):
    updates = {}
    if payload.title is not None:
        updates["title"] = payload.title
    if payload.period_start is not None:
        updates["period_start"] = payload.period_start.isoformat() if payload.period_start else None
    if payload.period_end is not None:
        updates["period_end"] = payload.period_end.isoformat() if payload.period_end else None
    if payload.meeting_date is not None:
        updates["meeting_date"] = payload.meeting_date.isoformat() if payload.meeting_date else None
    if payload.participants is not None:
        updates["participants"] = payload.participants
    if payload.summary is not None:
        updates["summary"] = payload.summary
    if payload.decisions is not None:
        updates["decisions"] = payload.decisions
    if payload.actions is not None:
        updates["actions"] = payload.actions
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    updates["updated_at"] = datetime.utcnow().isoformat()
    updates["updated_by"] = (_auth or {}).get("sub")
    updates["request_id"] = getattr(request.state, "request_id", None)
    updates["id"] = mr_id
    set_clause = ", ".join([f"{k} = :{k}" for k in updates.keys() if k != "id"])
    with engine.begin() as conn:
        res = conn.execute(text(f"UPDATE management_reviews SET {set_clause} WHERE id = :id"), updates)
        if res.rowcount == 0:
            raise HTTPException(status_code=404, detail="Management review not found")
        row = conn.execute(text("SELECT * FROM management_reviews WHERE id = :id"), {"id": mr_id}).mappings().first()
    return _row_to_mr(row)


# --- Exports ---
def _rows_to_csv(rows, headers):
    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(h) for h in headers])
    return sio.getvalue()


@app.get("/export/audits.csv")
def export_audits_csv(status: Optional[AuditStatus] = None, scheduled_from: Optional[date] = None, scheduled_to: Optional[date] = None, completed_from: Optional[date] = None, completed_to: Optional[date] = None, request: Request = None):
    rate_limit(f"export:{request.client.host if request and request.client else 'local'}")
    base = "SELECT id, title, description, status, scheduled_date, completed_date, created_at, updated_at FROM audits"
    filters = []
    params: dict = {}
    if status:
        filters.append("status = :status")
        params["status"] = status.value if isinstance(status, AuditStatus) else status
    if scheduled_from:
        filters.append("scheduled_date >= :sfrom")
        params["sfrom"] = scheduled_from.isoformat()
    if scheduled_to:
        filters.append("scheduled_date <= :sto")
        params["sto"] = scheduled_to.isoformat()
    if completed_from:
        filters.append("completed_date >= :cfrom")
        params["cfrom"] = completed_from.isoformat()
    if completed_to:
        filters.append("completed_date <= :cto")
        params["cto"] = completed_to.isoformat()
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY id"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
        data = [dict(r) for r in rows]
    csv_text = _rows_to_csv(data, ["id", "title", "description", "status", "scheduled_date", "completed_date", "created_at", "updated_at"])
    return Response(content=csv_text, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=audits.csv"})


@app.get("/export/nonconformities.csv")
def export_ncs_csv(status: Optional[NCStatusEnum] = None, severity: Optional[SeverityEnum] = None, clause_id: Optional[str] = None, audit_id: Optional[int] = None, created_from: Optional[date] = None, created_to: Optional[date] = None, due_from: Optional[date] = None, due_to: Optional[date] = None, request: Request = None):
    rate_limit(f"export:{request.client.host if request and request.client else 'local'}")
    base = "SELECT id, audit_id, clause_id, severity, status, description, corrective_action, owner, due_date, closed_date, created_at, updated_at FROM nonconformities"
    filters = []
    params: dict = {}
    if status:
        filters.append("status = :status")
        params["status"] = status.value if isinstance(status, NCStatusEnum) else status
    if severity:
        filters.append("severity = :severity")
        params["severity"] = severity.value if isinstance(severity, SeverityEnum) else severity
    if clause_id:
        filters.append("clause_id = :cid")
        params["cid"] = clause_id
    if audit_id is not None:
        filters.append("audit_id = :aid")
        params["aid"] = audit_id
    if created_from:
        filters.append("created_at >= :cfrom")
        params["cfrom"] = created_from.isoformat()
    if created_to:
        filters.append("created_at <= :cto")
        params["cto"] = created_to.isoformat()
    if due_from:
        filters.append("due_date >= :dfrom")
        params["dfrom"] = due_from.isoformat()
    if due_to:
        filters.append("due_date <= :dto")
        params["dto"] = due_to.isoformat()
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY id"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
        data = [dict(r) for r in rows]
    csv_text = _rows_to_csv(data, ["id", "audit_id", "clause_id", "severity", "status", "description", "corrective_action", "owner", "due_date", "closed_date", "created_at", "updated_at"])
    return Response(content=csv_text, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=nonconformities.csv"})


@app.get("/export/audits.xlsx")
def export_audits_xlsx(status: Optional[AuditStatus] = None, request: Request = None):
    rate_limit(f"export:{request.client.host if request and request.client else 'local'}")
    from openpyxl import Workbook

    base = "SELECT id, title, description, status, scheduled_date, completed_date, created_at, updated_at FROM audits"
    params: dict = {}
    if status:
        base += " WHERE status = :status"
        params["status"] = status.value if isinstance(status, AuditStatus) else status
    base += " ORDER BY id"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "audits"
    headers = ["id", "title", "description", "status", "scheduled_date", "completed_date", "created_at", "updated_at"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(content=bio.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=audits.xlsx"})


@app.get("/export/nonconformities.xlsx")
def export_ncs_xlsx(status: Optional[NCStatusEnum] = None, severity: Optional[SeverityEnum] = None, request: Request = None):
    rate_limit(f"export:{request.client.host if request and request.client else 'local'}")
    from openpyxl import Workbook

    base = "SELECT id, audit_id, clause_id, severity, status, description, corrective_action, owner, due_date, closed_date, created_at, updated_at FROM nonconformities"
    filters = []
    params: dict = {}
    if status:
        filters.append("status = :status")
        params["status"] = status.value if isinstance(status, NCStatusEnum) else status
    if severity:
        filters.append("severity = :severity")
        params["severity"] = severity.value if isinstance(severity, SeverityEnum) else severity
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY id"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "nonconformities"
    headers = ["id", "audit_id", "clause_id", "severity", "status", "description", "corrective_action", "owner", "due_date", "closed_date", "created_at", "updated_at"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(content=bio.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=nonconformities.xlsx"})


@app.get("/export/assessments.csv")
def export_assessments_csv(clause_id: Optional[str] = None, status: Optional[StatusEnum] = None, owner: Optional[str] = None, created_from: Optional[date] = None, created_to: Optional[date] = None, due_from: Optional[date] = None, due_to: Optional[date] = None, request: Request = None):
    rate_limit(f"export:{request.client.host if request and request.client else 'local'}")
    base = "SELECT id, clause_id, status, evidence, owner, due_date, created_at, updated_at FROM assessments"
    filters = []
    params: dict = {}
    if clause_id:
        filters.append("clause_id = :cid")
        params["cid"] = clause_id
    if status:
        filters.append("status = :status")
        params["status"] = status.value if isinstance(status, StatusEnum) else status
    if owner:
        filters.append("owner = :owner")
        params["owner"] = owner
    if created_from:
        filters.append("created_at >= :cfrom")
        params["cfrom"] = created_from.isoformat()
    if created_to:
        filters.append("created_at <= :cto")
        params["cto"] = created_to.isoformat()
    if due_from:
        filters.append("due_date >= :dfrom")
        params["dfrom"] = due_from.isoformat()
    if due_to:
        filters.append("due_date <= :dto")
        params["dto"] = due_to.isoformat()
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY id"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
    csv_text = _rows_to_csv(rows, ["id", "clause_id", "status", "evidence", "owner", "due_date", "created_at", "updated_at"])
    return Response(content=csv_text, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=assessments.csv"})


@app.get("/export/assessments.xlsx")
def export_assessments_xlsx(status: Optional[StatusEnum] = None, request: Request = None):
    rate_limit(f"export:{request.client.host if request and request.client else 'local'}")
    from openpyxl import Workbook
    base = "SELECT id, clause_id, status, evidence, owner, due_date, created_at, updated_at FROM assessments"
    params: dict = {}
    if status:
        base += " WHERE status = :status"
        params["status"] = status.value if isinstance(status, StatusEnum) else status
    base += " ORDER BY id"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "assessments"
    headers = ["id", "clause_id", "status", "evidence", "owner", "due_date", "created_at", "updated_at"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(content=bio.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=assessments.xlsx"})


@app.get("/export/audit_nonconformities.csv")
def export_audit_nc_csv(status: Optional[NCStatusEnum] = None, severity: Optional[SeverityEnum] = None, audit_id: Optional[int] = None, request: Request = None):
    rate_limit(f"export:{request.client.host if request and request.client else 'local'}")
    base = """
    SELECT a.id as audit_id, a.title as audit_title, n.id as nc_id, n.severity, n.status, n.description, n.owner, n.due_date, n.closed_date, n.created_at
    FROM audits a LEFT JOIN nonconformities n ON n.audit_id = a.id
    """
    filters = []
    params: dict = {}
    if audit_id is not None:
        filters.append("a.id = :aid")
        params["aid"] = audit_id
    if status:
        filters.append("n.status = :status")
        params["status"] = status.value if isinstance(status, NCStatusEnum) else status
    if severity:
        filters.append("n.severity = :severity")
        params["severity"] = severity.value if isinstance(severity, SeverityEnum) else severity
    if filters:
        base += " WHERE " + " AND ".join(filters)
    base += " ORDER BY a.id, n.id"
    with engine.connect() as conn:
        rows = conn.execute(text(base), params).mappings().all()
    headers = ["audit_id", "audit_title", "nc_id", "severity", "status", "description", "owner", "due_date", "closed_date", "created_at"]
    csv_text = _rows_to_csv(rows, headers)
    return Response(content=csv_text, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=audit_nonconformities.csv"})


@app.get("/export/audit_nonconformities.xlsx")
def export_audit_nc_xlsx(request: Request = None):
    rate_limit(f"export:{request.client.host if request and request.client else 'local'}")
    from openpyxl import Workbook
    with engine.connect() as conn:
        rows = conn.execute(text(
            "SELECT a.id as audit_id, a.title as audit_title, n.id as nc_id, n.severity, n.status, n.description, n.owner, n.due_date, n.closed_date, n.created_at FROM audits a LEFT JOIN nonconformities n ON n.audit_id = a.id ORDER BY a.id, n.id"
        )).mappings().all()
    headers = ["audit_id", "audit_title", "nc_id", "severity", "status", "description", "owner", "due_date", "closed_date", "created_at"]
    wb = Workbook()
    ws = wb.active
    ws.title = "audit_ncs"
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(content=bio.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=audit_nonconformities.xlsx"})


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=False)
    # org scoping (dev only if org present)
    org_id = getattr(request.state, "org_id", None) if 'request' in locals() else None
    if org_id:
        filters.append("org_id = :org_id")
        params["org_id"] = org_id
